__author__ = "Nessa Carson"
__copyright__ = "Copyright 2020"
__version__ = "1.1"
__email__ = "nessa.carson@syngenta.com"
__status__ = "Production"

##import glob
##import html
import os
##import urllib.request
##import urllib.error
##import string
import threading
import time
##
##import openpyxl
##from openpyxl.styles import Alignment, Font
import PySimpleGUI as sg
import pythoncom

def importPackages():
    global packagesloaded
    global glob, html, urllib, string, threading, time
    global openpyxl, Alignment, Font, pythoncom
    pythoncom.CoInitialize()
    
    import glob
    import html
##    import os
    import urllib.request
    import urllib.error
    import string
##    import threading
##    import time

    import openpyxl
    from openpyxl.styles import Alignment, Font
##    import PySimpleGUI as sg
##    import pythoncom

    packagesloaded = True

def getevent(position):
    for n, line in enumerate(page[position:]):
        if line.strip().startswith('</tr>'):
            endposition = position + n + 1
            break
    event = [p.strip() for p in page[position:endposition]]

    return event

def captureline(line):
    result = line.partition('">')[2].partition('</td>')[0]

    return result

def decodethemes(themelist):
    replacelist = [('agro', 'agrochemistry'), ('anal', 'analytical'), ('chembio', 'chem bio'), ('comp', 'computational/data'), ('edu', 'education'), ('inorg', 'inorganic/materials'), ('medchem', 'med chem'), ('policy', 'law/policy'), ('pharm', 'pharma/regulatory')]
    otherslist = ['automation', 'careers', 'diversity', 'process', 'synthesis']

    if 'all' in themelist:
        outputlist = sorted([r[1] for r in replacelist] + otherslist)
    else:
        names = [r[0] for r in replacelist]
        newthemelist = []

        for theme in themelist:
            if theme in names:
                newthemelist.append(replacelist[names.index(theme)][1])
                if theme == 'chembio':
                    newthemelist.append('med chem')
                if theme in ['medchem', 'process']:
                    newthemelist.append('synthesis')
            else:
                newthemelist.append(theme)

        outputlist = [i for n, i in enumerate(newthemelist) if i not in newthemelist[:n]]
        outputlist.sort()

    themes = (', ').join(outputlist)

    return themes

def decoderegion(region):
    continents = [('NA', 'North America'), ('SA', 'South America'), ('Aus', 'Australasia')]
    Eurlocalities = [('WEur', 'Western Europe'), ('EEur', 'Eastern Europe')]
    USAlocalities = [('WUSA', 'Western USA'), ('EUSA', 'Eastern USA'), ('CUSA', 'Central USA')]
    UKlocalities = [('NEng', 'Northern England'), ('SEng', 'Southern England'), ('NI', 'Northern Ireland')]
    regionlist = continents + Eurlocalities + USAlocalities + UKlocalities

    if region in [r[0] for r in regionlist]:
        output = regionlist[[r[0] for r in regionlist].index(region)][1]
    else:
        output = region

    return output

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.environ.get('_MEIPASS2', os.path.abspath('.'))
    newpath = os.path.join(base_path, relative_path)

    return newpath

def getfile(message, **kwargs):

    if 'defaultvalue' in kwargs:
        defaultvalue = kwargs['defaultvalue']
    else:
        defaultvalue = ''

    if conferencelistexists:
        noFileButtonText = 'Overwrite Conference_List.xlsx'
    else:
        noFileButtonText = 'Create new file'

    layout = [  [sg.Text(message)],
                [sg.InputText('_', key = 'inputfile'), sg.FileBrowse(initial_folder = mydir, file_types = (('xlsx files', '*.xlsx'),))],
                [sg.Button('Add to file', key = 'addfile', bind_return_key = True), sg.Button(noFileButtonText, key = 'nofile')]
                ]

    window = sg.Window(f'Syngenta Conference List v.{__version__}', layout, return_keyboard_events = True, icon = resource_path('syngenta_logo.ico'))
    window.Finalize()
    window['inputfile'].update(defaultvalue)
    window.TKroot.focus_force()
    window.Element('inputfile').SetFocus()

    while True:
        try:
            event, values = window.Read()
            if event == 'addfile':
                inputfile = values['inputfile']
                break
            elif event == 'nofile':
                inputfile = None
                break
            elif event == sg.WIN_CLOSED:
                window.close()
                break
        except KeyboardInterrupt:
                window.close()
    window.close()

    return inputfile # Throws exception if window had been closed unexpectedly

def getDataAndInputFile():
    global datafound, page, errors, lod, outlist, potentiallist, conferencelistexists
    pythoncom.CoInitialize()
    
    # Scrape from the web
    url = 'http://supersciencegrl.co.uk/conferences'
    try:
        with urllib.request.urlopen(url) as response:
            htmlr = response.readlines()
    except urllib.error.HTTPError as error:
        sg.Popup(f'{error.code}\n{error.read}')
        print(error.code, error.read)

    page = []
    errors = []
    for h in htmlr:
        try:
            page.append(h.decode('utf-8'))
        except UnicodeDecodeError:
            errors.append(h)

    lod = []
    for n, p in enumerate(page):
        if p.strip().startswith('<tr class="body') and not 'postponed' in p and not 'cancelled' in p:
            event = getevent(n)
            mydict = {}
            ColumnOneEnd = False

            for n, line in [(n, line.strip()) for n, line in enumerate(event)]:
                if n == 0:
                    themelist = []
                    for kw in line.split(' '):
                        if kw.startswith('c') and not kw.startswith('class'):
                            themelist.append(kw[1:])
                        elif kw.startswith('l'):
                            region = kw.translate(str.maketrans('', '', string.punctuation)).strip()[1:]
                            region = decoderegion(region)
                            mydict['region'] = region
                    mydict['themes'] = decodethemes(themelist)
                elif n == 1:
                    mydict['url'] = line.partition('href="')[2].partition('"')[0]
                    if line.endswith('</td>'):
                        title = line.partition('rel = "noopener">')[2].partition('</td>')[0]
                        title = title.replace('<span class="new-fa"><i class="fa fa-certificate" aria-hidden="true"></i></span> ', '')
                        mydict['title'] = title
                        ColumnOneEnd = True
                elif line.startswith('<span class="tooltipconf">') or line.startswith('<!--<span class="tooltipconf'):
                    pass
                elif not ColumnOneEnd and line.endswith('</td>'):
                    title = line.partition('</td>')[0]
                    title = title.replace('<span class="new-fa"><i class="fa fa-certificate" aria-hidden="true"></i></span> ', '').replace('</a>', '')
                    mydict['title'] = title
                    ColumnOneEnd = True
                elif line.startswith('<td class="column2'):
                    mydict['startdate'] = captureline(line).replace('<nobr>', '').replace('</nobr>', '')
                elif line.startswith('<td class="column3'):
                    mydict['enddate'] = captureline(line).replace('<nobr>', '').replace('</nobr>', '')
                elif line.startswith('<td class="column4'):
                    mydict['location'] = captureline(line)
                elif line.startswith('<td class="column5'):
                    price = line.partition('>')[2].partition('</td>')[0]
                    price = price.replace('/', '').replace('<em>', '')
                    mydict['memberprice'] = html.unescape(price)
                elif line.startswith('<td class="column6'):
                    price = line.partition('>')[2].partition('</td>')[0]
                    price = price.replace('/', '').replace('<em>', '')
                    mydict['nonmemberprice'] = html.unescape(price)
            lod.append(mydict)

    outlist = []
    for event in lod:
        eventlist = []
        eventlist.append(event['title'])
        if event['enddate'] != '&mdash;':
            eventlist.append(f'{event["startdate"]}–{event["enddate"]}')
        else:
            eventlist.append(event['startdate'])
        eventlist.append(event['location'])
        eventlist += ['', '', '', ''] # Empty cells for Syngenta list
        if event['memberprice'] == event['nonmemberprice']:
            eventlist.append(event['memberprice'])
        else:
            eventlist.append(f'Members: {event["memberprice"]};\nNon-members: {event["nonmemberprice"]}')
        eventlist.append(event['url'])

        # Extra fields
        eventlist.append(event['themes'])
        eventlist.append(event['region'])
        outlist.append(eventlist)

    # Find default inputfile to append sheet
    xlsxlist = glob.glob('**.xls*')
    shortlist = glob.glob('**onference**.xls*')
    if 'Conference_List.xlsx' in shortlist:
        conferencelistexists = True
        shortlist.remove('Conference_List.xlsx')
    if not shortlist:
        potentiallist = xlsxlist
    else:
        potentiallist = shortlist

    datafound = True

def stopanimation():
    sg.PopupAnimated(None)

def permissionerrorpopup(error, filename):
    message = f'Permission denied: {filename} output file.'
    message = ('\n  ').join(message[i:i+64] for i in range(0, len(message), 64)) # Line breaks
    message += '\nPlease close the file to continue.'
    
    layout = [  [sg.Text(message)],
                [sg.Text('', size = (3, None)), sg.Button('OK', bind_return_key = True), sg.Text('', size = (30, None))]
                ]

    window = sg.Window(type(error).__name__, layout, keep_on_top = True, return_keyboard_events = True, right_click_menu = right)
    window.Finalize()
    window.TKroot.focus_force()

    print(error)
    while True:
        try:
            event, values = window.Read()
            if event in ['OK', sg.WIN_CLOSED, 'Escape:27', chr(13)]:
                break
            elif event == 'Exit':
                window.close()
                raise CloseAllWindows
        except (KeyboardInterrupt, CloseAllWindows):
            closeallwindows()
            return True # Error thrown
    window.close()

    return False # Error not thrown

def showreport():
    if len(outputfile) >= 30:
        textsize = (60, None)
    else:
        textsize = (None, None)
    
    layout = [  [sg.Text(f'{len(outlist)} files added to {outputfile}.', size = textsize)],
                [sg.Button('OK', bind_return_key = True), sg.Button('Open file', key = 'openfile')]
                ]

    window = sg.Window('Successful!', layout, return_keyboard_events = True, icon = resource_path('syngenta_logo.ico'))
    window.Finalize()

    while True:
        try:
            event, values = window.Read()
            if event == 'openfile':
                openfile = True
                break
            elif event == 'OK':
                openfile = False
                break
            elif event == sg.WIN_CLOSED:
                window.close()
                break
        except KeyboardInterrupt:
            break
    window.close()

    if openfile:
        openoutput()

def openoutput():
    os.system(f'start excel "{outputfile}"')

mydir = os.getcwd()
sg.theme('Green')
right = ['right', ['Exit']]
ring_blue = b'R0lGODlhQABAAOcAAAQFCASE6oyKjAxEaDyu9ExUXESCnAxipAQjPiSb7oTO5GyvxiREVwRksUSazCQmKCyCt8jO1AQyVGTH9wRMhip6sgQSGtzt+URteWSbqTRifGS85AyQ9hF1xCSM2WyGnBQjLCRSbCmm/KzG3AQ0XFSy6gRWlBgYHGx2fBSE2BRSfP///3S/1ZSqvESe3CxypAQaJAR0zxxDWzSQzMTi/HTT/ITc+P////T19Vi66hRXhwQqRDxOXDSCtBwyOUx4hxwsMzRTXCRKXHTK6GmnvAR61AQMEgSK9RxupBRrr////zSV1xY+XP///3S796yqrApKekRmbESGpHS2zFqnxwQVJmTC/ARYoByK2ESm5ODg4Bw7TpGdp0yx7P///yQ4PxRShOr0/GjC6hyO5zyn6v///////zR6pP///1R+hAyK7FSEkv///2yx5wRtwdTU2ARSlDR+rKza/FSy/BpKcIzj/DSi9QQsTzRdcP///wwODwyD4f///yRllCxGVlyarKy6zJC83LzQ4HyWrKSkqDw+RHimzMDBxDQyNKTQ9CxulP///3yIkMTa7CSW7CRWhJS21NTk9JSWnJSmtJTL9xx0tjx2jPr8/Ozu7lSg4FSKtFSq7Hx8fCSKzFzC/DmKvP///1SWrFSixDxuhLze/GSiuExaZP///zRqfP///0yWuczm/Dyi5P///3zO53zC/ARep9/m7CQ+RGS23JTC7IzK/CiGyyd+wle+/DSGxP///////yROaWym1FySvP///wQ8bVh8nP///0S2/HSOpBQeJNza3Gy21P///3yOnAwkNSyd8GzI9DxldGy63BSQ7CyO1xwmMSxTYv///wxXjBxRd3zC1Eyi7Mze7HzT9HzG3Hy67Hyy3P///yw6TDx7nDxebMzW3Ey29ByW7CRqlDRynwxSiDyi8TRuhAxenESx+VRSVP///xRmoAw1UQxOhAwUGeTu9HRydByF1wwaIwx2zDyQxIzc9QwrQTyGsCxLVAx7zyRwohxqrDyS3P7+/iH/C05FVFNDQVBFMi4wAwEAAAAh+QQJCQD/ACwAAAAAQABAAAAI/gD/CRxIsKDBgWESvZpjR4TDh+fmvJJz4aDFixgxXqLhRAQBXDXqiKxBsoaNkbgIiHBC41LGlzAFhqllR13ICcNEcDjCsydPDh4n3KuhTkStMDGTFgzzapmnOp5E+JxKtacIXPesLHuFVCnMS5SWTbDRhUOAqmirEsg2YRkll14vXjgn7p66s2eP5OWZt+9UvHzXziFTMa7BRAlcWVETAG9jvc/slJiQLdu9ezayMctBYFxjx49zuEqQyPDAS07IZLPzufWzLkNsiOlix9GzZ2rGjLPTZUPlHAk+622cIBsZJ3C9XrpWYgjj1gESiMmWw9GYTZRIXcBx6RKOC6Ro/s3x4KiEgiGsoe8RU2JT8piXsuRwtidAEfsBnm1wxQqLk1UxreKEGqwMMYQj97WWQw7XvJfRciU401iCRWShQBYevBUXDpR4wIoCOexxXxH3bTJLgzG1UcIsJNpXxB4bOIPFNjiYJhAO26QwCwtjjEhiCVk4A1MiS7DQYgxFYMFCFtAAaCNBq0CThTZLxBAAkkUcs0RpGEXiAQspkIjkPCwskYmDT/6DQyYeUIkliSx4EMlFl/jjDDQx5BlDCixA0waaaf5zSRseWINnnkXMc4w/gP5DiwNU6BnDPlMs0UagGPXSCQu2SOqAA7QcFIYtx9QjKRUOnIkpnZnMcMw+/jG4kecCtnRFEDeizODGrm7YQ4UtNa56EQ62iEIFrzEUy81StxDBqxsdLFDJnMJ26QYRuTxbyi22/rONKp/w2oAonyxbbUbcQODsrg3coso2pzXrRgP0VlJKJY2ee1oHpfRA77zbwtVIHFLQS68qPYSqb0aBVFDKvPT2AEEjAvUSSiUGJ/EHvgtnhIPGFcBCbxKh9CJQBRnAojIsPeRjSMcv9RKHJivDEkoF/8TDjxRXqHxFKEhgA3NG2ByQwQErx8FPPCOc8cIVPR+wBhJDv9SPJorAYsIV/JQzgiEGtHPF1uQY4EvVGflyhiVbm5COAYZosoYJdJtQjiKQoI0R/iTkzF33GpqUswYchMNhSR+C6H2RIO0MTjgFBpTTxw8UUAAHBT9QE4viFsViThqXX354H5ZUXnnmwXK+FAVpUGM6OjqAgYrpmFOQr96XYG5O5VBooIMKGkABDAXA/ADF7WhfAgUGKghPgQZgPBL88MBgAEW3qt9YPRTC965CNc0AIzwwzUCxefYExbL8AOIDg4f04QMjPx50JI7+QIIMAA778/cRDB7yk18IeNGC+w2kBfoLoPuCMQhp8I8EdAhBMAwokDVIgxckAEYGpTEIQMhABiQIIROkQQcK/oMX0gBhCAfAAEDEggkqJIEEguCOcBgwHCSgYQhJIAMmbI4X/kG4gwQkcIcPEsOAxNiCPu5AAiYygBcC+QAD3HEHIbrDD0xIHedwwAQGMKGKQmTABwQSAXdsAYx3kIU7JoG+SbhDFvgAozvcEQGBXIIBskDAHfToji9kUXVc/MIZd0BIWfghOVzwAT4QwEgEfMEHyVBdMvDhjR0wcgf48MEgCIKJPjYSATtQpA31Fo5M+uCTPnAHJgrCCHwsspGZ9IMWO4YDP+DjC59EABAYYRBM+AAICIABI2EABGVEAXlpukQUQOADZQhTmEDAxyoNwgVlgAAG2GQkEECAgqFxQpfOFGYVlKEMLljkEjwAAQKqAIMqjBME9PgAMuNyiQ/AAATX/mwnO0HAg3wZw5rsdOc7YWCKWaYJB6agBz4F2k4QKMMYGSEEDOghUHfCIBr08EYdMfUGbyiUohV1JhtfggJ6gFSgFjApDBhhUKXggBEmLUZFBUqPbn4lofSwgE53Co8TVAEIXGipx7gAhJTCowoWMMJOq0CPAszzRgWwADyUqlSdGgEe8LAACDix0ZdEgBPFuOoJrJrUKhihCvAogFCHFdWpGuGtVJWqHt4KBFNI4hBawER3MKGFQ0giCg94K1aTCtfCWkCthrkECgrL2LfK1a2Nfase9ADZxhJWHk+1CCHcCgAjdPaznoXrXCVL2tHCFbRwhQchMGUMb7wVALANZO1nQTtbz8YWtrT9bCEguqpLSOIEs8WtcIdL3OICgB6SyGxSMMEJPRj3uc+FByem2TFMMOIB0M0uAB7ACOpW7RKHkEc0tCvcB8jjEMqtlhYIIY91IOIEzoXtCRCxDnkQQguBCggAIfkECQkA/wAsAAAAAEAAQACHBAkOBIfvjIqMBEh/RLL2TFFU////DGisCic5jMr8SZ/RKUdQJCYsKJ/1xMbMBGi4ZLbU////////ZMb4ZKCw0eb5DBYbBFCPRG50bImcJHe2LFdvbLPpETZM////b3R3////JIzYFBkaBHbMSJLGBDxpkazBFygrpMrszNrk7vDxFFSE////////BFyjFIfgZMDuNJ/pRH6cJ1h+////////////////JDlEBHjWrKqs////3N3gTJS0NFBc////////Bw8U////hNftdsDXfMjhbKPPCx8pbJKsMHCaFDBBNzc5FElsHGieCTBLWavP5OjocYKR////HI/jXLPc////PJTMVG6EdLTsRJrMBEF1rNb4VK33////+vz8////NFp8NpjaVF5oVJ7EJE9mBHDIdLLE////////OH60O2FxPJLcFCAnHFqBeb/5PKboJEBMHILU////4O/8VLTuDI70lJyh////////RKrkPHKM////////DEJopKq0////rLrMvMjQbK7EfKbMVIiYfKTEVIKMfJKsv8LEhNDlxNLgpMLcpNL8VHaU////PHaXXGZs////lLXR////////PEJM////JDI6vNLkFFaMfIeRHDhGfHp8vNvxNIzJpKKkXI60KYDAi9/5PFhgPGd4lM78j5OUVFhYZK7MVI68XJWmfH+H1ODwXJ7U1NLUaLrc////HE5wHG6oT2Z8dLrRdK7gVJjYBE6E////aLrsb3qE8vb4////V67gLH6sFEJk////GmKYaJq0ebrsPI68////////KEhk////LHKo////////LGKQTKbUOKb4PGKEBFaYrLK8XLrkfLLcVLr4lKKspLLEXIKkfI6cHD5MDIjqDEp0FGaglMjwbLLIbMnxTGqENFhkLI7ZDDxcrMbc1NfcHFR9DFycHIfWbMLp////THyMXKbkLDhIVJOsPE5cHDE6HEhkJGiWXGp0fLTkTJe/XJ+8LFFiDHHBHCAkLEBI////zNHUrM7sXHKE/v7+CP4A/wkcSLCgwYFztrjh4qyBw4cxuAzbMuegxYsYMXrp5KaBM2rfRIka8q3kt0RDEn2j0yAMh05eMsqcKXBXAmcEQn4j0KBOgJ9Af2qb8kZaonRv1rjZRbNpwV0dJ4ia0DOAz6A/c2QNoDXHlF5FeoUYxtTpTC+lGkglcBVrUK1c4+bomiMGkSchusU0e3HOR1EE3GLtKpcr3blz17zKsqYiX4OMGgyZ0BZrnRh0YJxMxJlIrmYx0CFGPKK0Am+hUDwe6MWNM1ENBNchUGQIjDxhprx4oe3FlDBvnhQh0kx06TIjypSJg8pTrb1md3GhNqTyzwYwEkmbMoVDgk5zdv558aJiTqdu7OKgU0AElRXl8Mvgs1fPVtmm0jVbr5OuSJ4pWFRAUwVYxOEJKqigA98DZTxQTw8k3CeTF1zAMIFbeSTyXykS0rRLNxp44k09+DxgoonEQAhdRrnQ8c1b2qSTzhTDrMjXLvTgMwYFGpz4gAvEyGDETIw4kwhQWk2hYRgCrlYQK6EQQ0EoLvz4ozvJLJJRBQ0k4hNcSsbQTIdO1lSPBhT4UqULVbrTBBQXefFGETHINUURLtlYJmtGyMIjm2xyQ0gaegrUDRfSwJWDNnhysGdGfarSBKDQPJKEJAftoqQ2hgXwSh7skPkoQV6kkowqLkCjKjSEBKOCQf7DSJPHXFwp8EoIoo5K0C7JyCADNBcAG4wehTx1J10vFBFHk7pilMI5hshzwbQX6NHGqwO5QUUMo1GRxTDNzlRIMIZMe8sAbSRhAmtrFDHaFLTEkWu4uwazjjwDnDsABszs1UkYT4z2hBXd0DuTJG1goO8AM7ShiEBYQBDCXCPEEe+8Bg+0SyZ6zDDAxwOogYRA4hBx3AhWZFFLxjMJswEpWsSshRoz/DOHJ83Eh0oozLJ8UQrbYDBACVqUsAETUKCgwHvKjWBGHD7PZI4asZRgNROxAELPE3HAF8oYrUQt0zUbbGC11fcgYYs3JTLYAzGYio2RCUyYXYITJfjQiP4ngjBoYg+hYCI3RpiMM8rdd98DRigU+GiPLI4NfhAPJYziBN5OwBHP1ye6QAE+GMutghPh3H35L+No0IOVD1BwQKGS/+NFCQt0cLkT2XSAjztWuiDMObBL7oUTC9yOwCZOaOAOoC6ockDoYo9evBMIINBBB8oz704wcMZuECtO6EN99ZvAkQaqbEIjQxOCe19QIJlXX70Sx6SyfKqVXup+QdU4gUP1RzgCPGYxCBkEY1XyeAQo9kcQfyhBCQFEwBGUoIlFPEIewLoAtILBwIFkAx4nkGAEozGHYDyCWrdYxwpS0EF+IAAeAYzhEVjxjySs4xYXOJc8SCGMDkYBAf4nCKAFAggHgQhDDyvQ1wowkAnoGcwLSgjhEYY4RF0IRBEzwBfISGGOuHnPDkcIogXGGEAHCMQLzBDax7RgDi46sVm7gAcb2DBFMioBOiZghjlkNrNYHMJ7mgjjGIMwRguYgiAqWAHMilYCJqihDywcXDmOQMdBEvII2BrIITZQtbOR4R4zeGOZdtGOI4jAAkEg5BitWBAVmCMcfThbCcLxi0YEb1ReEIMF2EDIVF4ykwQxARmqhrlx+KADGYjaBwqJSl8GwQ4W8cIz7jGO2znBmMi85Wq8oAsqOjOVlQgeD35xD8RdrgPFu4IoZ7ILMfTSlwAIAgAsUI6MWAOd1v68nD7oV89RuaIdqkxlPAFA0GrMJAPZyMb4EHA5HGwCAdhYp0F2sQpUNpOgGA3CB2jiBXBsogPyk58S/peNakj0H7uwwwlSiUqMurQA2tzVLD4KwJCCEAFKiAI/aMKPD+QjoC516Tti+pRZKAEeIpQhEIMowHnYARE8UMF4VMADRNgBEitVZVCDCtPHeOGHSkgqFQM4RzrW8Z3OtORWXfoBol7ED0BEwFgLWccp2pWZal2rSw36KB744AgSxGsvm9lMZ+o1qEvoJy5VKkSLsvSb8pTnYTNqCrc6RQWrkOsUIRvZyQaVE91jmQpMAQ9mClSyngVAPgQQWrF5ARG6gCRHPDs72Xx8ABGWpRcU/KALMbQjHxYgqAVE0I4CcOITPHhUQAAAIfkECQkA/wAsAAAAAEAAQACHBBEdFIzkBFCNjI6UTE5UVI2fCm66Bi9MXLLUJE5klMr0////JDE5RKPcVKLEdMnnNIq8DD5gBB8zzNTYRHB8BH/hJF6EBFmfbLjUbKLMb3V8JHS43Ov3jLLUFCAmNJXcBHDIbIykBDBWRH+cbL7c////BEBxrK+3HFqBJEBKOmBqKZfpLHqs////BBYkBGSyBCI8bLPqTK/nRJrMzODw9PX4FC9CZKrBlL/h////HE9xZL/p////////XG58FIHRJILEOUFJJIzY////NG+MbK3AbJe4HG6pxMzURoyxHEFXFicrPJvXB3bLFBocFDZMZLXZJGiSVIWW////////////////ZKC0gtHqTHefdIGP6/T5FjpU////JEZUfLXnKZLfjKK0OGqQBChG////Gne6SHeIfIyYBTZb////JFd3FSo2XLrpDCY0pLLEfMDUWK7cPGd3DBggHDA7pMLcTKj0GYbZJIbRPHakfK3UwNvvRJ3W////VGFs////JDhAvMbOpKSkrMzoRGZ0VHqcPH2fjJqskrjZ////vNbslKe3XJ7U3NrcfKbITIKsNldi+vz8hMrkJ2KR7O7vdKrYfJevLHapPJLE////////////ZJqsVGp8VIKofL70HEZk////LFZkXKbE1NrkBF6nbKa8TIac////NIKsdLr0TJ7E1Ob0LIbENEZUPHaURJK4////3Ob0fIaUxOL8qNL0WLXoLHKoNI3JbMLpGVyP////////////////VJbU////OJ7qBDllVKr0DIbkbHqEJ3q3RKrsWKbMDDZMrLbEfLrvfJKkpLrQVIq0DBIULFBeTKTXXKG/fMjfFEBeDB8pTHJ8dLnNlK7EPJbUdIqcTICS////NHykdLPp////1N7pHILNLIC8LI3XPG+KHDdF////fICGLEhRlKCpDCg+LFhxLITMXHiQTFZcZLrcZKa4LF54xMbIlJqcDF6cTJKs5ObnHIrcPIu31NbUDIHeDFqcdKDELHa05O75dMDY/v7+CP4A/wkcSLCgwYH9aCkT9gGMkIdCxH1Y1E1Qv4MYM2rUCGlWtxXAZLiTJs0fBgQIoGCwdnIGq3WU9EDaSLOmwC0KVhjbgQUBMDD5KjQR2qQJCBBNwEFQVUQUhHCHtticWnCLpxW1sNQCU6GrUKJFkR4da6DMq3evNuSRSrUmpJw8jXntOrSoXaNjQRgA8YLvi3DRkmw4NLNtRg4hscid+7Wu3bxHXxh4QZkyKRabuKGKZfggrRUPdgxj/DUfGGhwoHSzxrKIg1cQyvStbJlUoQJH6HQeCCkVMCzASFf40QDDGwQzxIH70aRM83WX6N24kkT2C1LYL1w4IiXKvsJtIf7VqfUgwOjRXYUgkAZNCLhFChJxqAEJUg0OiXAsKlPm3iZ6lmBHinYX6DMCEc2AZxMkxrCByzDodWVHLf4wYUc3NNhEAyVHWEIPPVEQeIEAF7hChCMKbiSeg6Q1IM0eQihQg2E14HCELQWMoB2JAghAxIkpZhRDLbhUEGE+O0BhhzIz7vZPDXnMM0IBufRopRiSGFETLcA8EGEF9jwAjTirOFnQKrZE0Z2VApgwDgrMbLTKClgEMFcA0nwATZNmElRDM7loI0mbPZpAgQ5lYgTJbytEaI80YHQTZJ//QGJELmZY0KYJPcYhxqT/KCBDLXPl488HMVCqkREomKHDpv4mqKHGNQdtQeeXUEADDaiqVtqJBRRwasKwKujAFkHK1NKAV03sAYU4fPZ6UA2SwBPHsMN+okYlVa0gDbN2PABOotJmRIMAg+gwbDDBPPLJJAR5oixdTdQyQzfl0lSJDiqYwG4wESQQBm8ffFuXPdbYEW2+0+oATwL/BvNMAoXNwgSpdcGBjQIM03TNJ4+wi4YISkyDhEAxuCMEUfm8AQ6vHQtUgw6hfILGzWicE4JAYETi2B6qpBrzRiF88owISIvgxTP/9AMGHI9BcUeGQ2s0QQSPJC0CF0/EQgs0TNiVjzXgVE3TJ+dMg/QYyJDjxhcI2IOXONEsYvZGhHihxP4YIowxhhJnQPPGYzNgg8PdGinCRQp9+/0HJx9YIxYIDrCiB+IZAYLMOX77Tc4z4mCQ1w1lcIA5RrGMkULnY8zBxR035FVEGceeTtAWY3DuNww22ACOKGO9UIQBMN8NierpjAEDDGukA84xevU1fPFmH/+H8hJIsEYbQIgyGV+lGLCw7f9sAcMfy8OgfTrrvDPbC9EcYTr5BDEyBvrqS7AEObfEXhk9xUgE/QgCiHQwIHvZ88A5fPGOyVAmCag4xAAHgo50zCF7LpCAB/qQB1UcoTIsSIIvJigQHyxhCRLIoAs8kA1B3CMclTnCJo5Awn+QY3sYlIAcTtCPDSThOv4Dws0oJogEGKDQBUikhgvq8Q/MCMg2eMjABLXQhiWo0AXU+INAMkCPI1xgQPMoQBTGd7ca2GAJSkSiC+RADIHowRKFGJB2TBEFCZIPHdSwYgqRKId4CAQSliiAiKKgjVxQr1w1mIMH0ugCAMiBAeDpQCFCNCIBjCAKjbCdLFxgRSQCwAXOWMbtuFNJAQRKAN/AHD40mEYAOFIO8CJII1wRBTaNYxySIGO+apACD3jgk650pTkMsoVcmIJNAqCABTpxSCdBog9y8EAjg+kMD8SyIIcQQy0JpYNDaWloGlihJ4MJAHkoCg9mQIGwTICCdBmhmVSBBDFWSA1yAsAZrf7gVSxQQAFCDQsFKviENnS5mxr0YYVysKcjGbERZqDgWtgyAb8cNoFe4eMP0UyoQgNREyPAQxIRNcE0HhGKCFSCoDWpgSx06AGN2lMDC8qCCiyALXaZ4BmPmMYnFIFSjNQAHXOIpjQV6gwCwBMSWVCHOiIWDDRw4Rnn4MI0slHRmkxAC1b0pUvtGYSeGqQGWUhAKPyFs5Ep4RzneMIBlMAOdAAiFpOoTw0YAQh0+GAOWNTqNBVKAK8eBBJFe0QERDCypI3hCSlIwR/IcYA2LA+B2aNGHvW615f6NSNumMYzlGDYxvntCTZgwB/+wIA5mPaES1ikGhV6T2eYk1LfiInDWUfGN9alT335w6AaV6vQPzC0VzVQxCcSiwzbpg+BV0wuUV17WapMYhlcIEcKnqC8x0JWjSkEJjmd4UhzXLNjkzCEFyx4wzEgN4efbGQjncFdBgzgu1WDBBK04IU2rOG+qW2DHKixXzn4t73EiEdz81UPN8iiD+cgRxskS401BKEPWjgBZ/oUEAAh+QQJCQD/ACwAAAAAQABAAIcEER0UjOQEUI2MjpRMTlRUjZ8KbroGL0xUtOgsT1mUyvT///8kMTlEo9xUosQMPmR0yecEHzM0irxEb3zM1NgEf+EkXoRsuNQEWZ5woMRsdXwkeLTc6/eMstQUICYEcMgEMFY0lNpEf5xsjqxsvtRUoeIEQHGsr7ckP0g0XmwUXJD///8seqwEFyQkl+kEZLIUL0JMr+cEIjzM4PBss+n09fj///88lsQUeMgcUHJkv+k8pegUgdFsrb8kgsSMwvQ0PkxUmrw0b4xcqsxccHwWJytRgakUOEwcbqnEzNQHdss8m9cUGhwkjdhktdgcQVckaJIWOVT///////////8bZZ////9UlKxkn7GC0epGjLH////r8/krR1CMorQsV2partc4apAEKEZ8tOj///8VKjYad7pEdowFNltUhZZ8wdf///9EndYpkt5UYWxcuumkucx8jZy8xc48aHgMGCAcMDtMqPQbhtt8q9T///88dqTB2+////8MEhT///8kOECkpKSkwtyszOhEaHRUepx8oLw8fZ+MmqyTuNn///+81uyUp7fc2tz///98gYY2V2L6/PyEyuQnYpHs7u////98lqxUqvQsdqlapsj///////////9Uanx8vvQcRmR8psz///9IdoTU2uQEXqdsprxMhpw0gqzU5vRwuvQkWn////80RlQ8dpRMhrT////c5vREkrjE4vyo0vQscqg0jclswukkQlj///////9UltT///83nekEOWYMhuRveoQsfrREquwMNkystsRkprh8hpT///9MpNdcob8UQF58yOAMIClMcnx0uc2UrsRMgJJ0ipxcoNgcXIw0fKT////U3ul0s+ccgs2Uwec8b4ocN0Qsjdf///9clKeUoKkMKD3////ExshceJBMVlxkutx8uulMnsQsXniUmpwMXpx0pswsWnjk5ugcitxctug8i7fU1tQMgd4MWpwseLjk7vk8ltR0wNj///88YWpEmswcesREpeQshMI8QkhclsRkq8f+/v4I/gD/CRxIsKDBgfNkjSkRQpsPa9Z87KOVq1qgeQczaty4EdKeam3aGANTj9kFf0MwYfLXY1gQLRvkpdsDiaPNmwK5/GizY1wyMEuaKFHyoegHAy9eGEDySwsWLZcudeCCs2pBLp3aIMgCpkmFoUONHn3xIanZUUgMcTOEpBBVqzchKXDhswG8r1/DEj1aluzZUaMwYICSxlAVRDXhauSwK0aWBhUigw0r1q/ZpIADCx7sLIyQV4oPynKRrF2vyHjBwtOGzwHLHj1IHbvijgUSDJoxCNgtIMyZaHBCD4SEaleWXaiV5IXHxgkzTDf25TNjoDqSDaZacbtieDNvASom/qQakRguJDsIIARAnZcdGDXlHkK7pogDF0iQanBQhKgfkirSFOAMFN8JYIIAc5hjRA3mAaNDLXcl14sT5IRgTTUz4DRDBlVA4YwI0ex24IHmmHMGgzid94YO7EXGhhpL3HFNeVVxgUgVknRmoIEmmKDOF4TQuBEqCNTSIjxOOHHHGELCVcMnAkwwQQ4j9vhFDs/cJMsuEJx2WgXs1GOMNqcIZ9AMklhwRio99sjLF54EtxEHLmQRwJdgJhOCMU2a+U8NRuQwB5Um8FLoI8hQoxEkO0CwC57sJNNGNX36+Q8kIwjqSaG8oPFAAvdUqkAMb7AHDwQh0GApR5ne84AJ/mjE6skTixzERZ294DmOMSVUuqpAkBDiSQq8dAoCCF14MolBnegADHv6OKGNr7/mpE4CCcSKBghooBDHVXWy10skd5RZLUeimIDosceiEMWyA/3wxg7sObEEOefeVAkyCbArRjDZHDIcrqiFeQe1+QpUgyddPAGCGBCjYEtisezQTr1LKJDwTYs80IUYD4sBAwxJCERDLS6gxoMaB29sUw3IdBEFxBBn44tALkSSXAPGqOqyTc8cgQLNMmRjyz/z7FKqckrU0kSGP3NEwQFAQCyDDGV4s44sMeygXAXwqGFN1Dc9kU0wV8sgRhEnkDOOV1+1MQQ0ZNsUTjYwXB1B/gRlPGNJMncNZQw+19TNUTcH/BGBDHuX4UYIyeSlBBja7GH4RnJ4o/jeERTRRRv1MK2EE9ZwcLlGr8iweQQtlFFHG+NMVg0Pb51u0CQy1ME46xF40AQYFXxAVOgI1w2JDAzw3oIHEfwO1gfM4FA82TXkrrwydDThxF7QGzB91JN0rgzrLdAhgzYXCF+UP/mYbrtBr3ROfgvKMBACM+p/gMk+irxvkByda4EAyweEEjCDe7BwR+H8R5BzKKMIAwQAHdxQjSFYwygSCALdGDgQTnjAAwMsny8EgQ9aFOUFOBiGGTg4kDpAUIAAkOAJ5uGDchigL8NAAtQYSIEWvDCG/jFcxz+00YOylMUd7kgHC32hDA8AoAUxpAMDBJIOB+TDLGbAAhK+l68alMEDdIBiDPugAYHsQQKwUEpSrrABRDCwG8sTIxDBAaxfDOMyG+AGElB0uhrUAYxy7AMDyoMIWPwCM6MowCzw8D5ixBGIMRwAQbhghn5gBgNIKEAVdlg3CtABjJAEQB/gNRA8aIEFmcGAIQwxCz5GrQZ/aGIoAeAIg3Ahj4DZTBqE0AouhgYSbvgkAIZJTCaQkiCIkIYhNoMBFaShCiMgmy8eScxhSvIgkDBFAW6jGwxEIw0qIM/GIDFNUFYTAEDw1SuQkIZu7sYCzojGgrrohuUp45yi/nwHRwIBBRHohjepCEUqJKGoX1EACMujAz4BcI6bZEAIrPiOCcKTAgFUwpcGqQExItDEe+KzjHFpBSuEsKMRzWEQOchBM1yJkxp0ow69A+FCCcBFSIjAMyXtUQ7ukQJP5GAEFMAJBRxRht4VIYz47AMBjnkTQEkCGzxqkwly8IhHeOIBngjHIuTwihrgpwaMkEM3iJANjpahCCCUYzUJwNIUjSAVg5JqsXjBr0ck6wjBoJkY1CYGb5TBdUVQRgSr2QcAaKCtVoGDoNjkpmJtCw1ReEIXuoCCyv7hsgxw4fhaEIEnnrMPdGjoqk4RhlT0dK7bYtcBIHYAb3ijr7vbhFsI5ahQIDDiXJBoRg6+8AVkaMtfIbOa3pQXwfIt7xwYvckkKuEJbHlCtXpNG+fmN0A6WJcYTE0YFwrxhShMFhl5Fe5weRcBOmBPinHI7s8gkYQReOII2cgGCrJRh5GdtQhoBWMEGOCLJCS3WuuAQxw4kYAnwMAbMvCG0dzwjBMU1E8BAQAh+QQJCQD/ACwAAAAAQABAAIcECQ4Eh++MiowESID///9MUVT///8MZ6kKJzmMyvxJn9EkJigpR1Aon/YEaLjExsxktNT///////9kxvcMFhvR5vlkoLAEUI9EcX5viZwkd7Zss+wRNkwsV2////8kidQUGRr///9sc3cEdsxIksYWJyv///8EPGmkyuzM2uQUh+H////u8PEEXKNkwe3///8klOUUVIQkOEQ0n+knWX3///9Efpz///+sqqz///8Eedbc3eBMlLQ0UFx3v9f///8YXIwUMEH///83NzkHDxT///8caJ6E1+0LHylso88sZpQcj+MUSWwJMEtcrdF8yOHc6PRkmqxxe4T///88oPT///88lMx4s+ZMl78EQXWs1vhUtOxUnsR0ssT///////98v/b///8kQFBUXmgkT2YEcMj///////9UboQ4frQ7YXEUICZEquT6/Pw0Wnz////g7/x8e38MjvSUnKH///////////+kqrQMQmj///8slOI8pulUgoxUrdwEgORclaasusy8ytRsrsR8psxUiJh8pMT///+/wsSE0OXE0uA8coykwtyk0vz///88dpdcZmyUtc7///88QkwwcJr///8kMjcUVowcN0a82vFUuvdcntQ0jMmkoqRcjrQpgMCL3/n///88WGA8Z3hUmryUzvx8ipePk5RUWFhUjrzU4PBUq/fU0tRoutz///9kuuQcbqgcTnD///90utBPZnx0ruQEToT////y9vkcWoQsfqz///8UQmT///8aYppxgpH///9cuuQ8jrz///////8oSGT///////////8sYpQscqhMptQpjtk3mtxcxvw4pvg8YoRUpswEVpissrx8stx8goyUoqykssRcgqS80uQcPkxsuvxUmNAMiOwMSXQUZqCUyPBssshsyfM0WGQMPFysxtzU19wch9cMXJxswuccVX0sOEj///9MfIxUk6w8TlwcMTokaJYcSGTk6vBctOBcn7wsQEgsUWIMccFcanQcICRMretcruwMftzM0dSszuz+/v4I/gD/CRxIsKDBgXD8Xdm2yZMGDfcepkE1aBGcgxgzatTYBtOVZcuwQAMHLgkXLO3a/fnTzoYjI0osJGqzsaZNgbfA6GEGwQc0K+buORjqoIVRo97gOeKjCBmyQrduSi2YE4aTJ04+jCjDtQzRoUdbRLtANoYSDGrSFZI39WabbzDmPdnDb4SOrVwdeC3agm+0vxdqDRicTk0HWJBots0IhYo+RGx0SL5bBi/RvkcBkxU8OEsWWGqIPUu1+CAjGIiAcdPhR7Jdrvc2BcNCz4LtP4RcwjtX68LgAZ5PnCBjb5e10gPbaNtzpIGf56zt8iMBARyXYLk0HNh+IGmypeyU/sQAngWP8BPjenDIoLjtLVWZEMl53trPCHNOZClwyOUbtgq3tNEGC1BgA8kfQADhBgaipJPFeU00cQID2aAR1VS36ONCOAHQ15oK88gCFC2Y3JSCBemkI4oaTEAYYTbZzHKhTW2oMkE4fnRInwJPWPHBNzNKdQsksIBGxgkRItkEB+/M0t5G2mTCYYcdqsAKBCr4EORitxQyjj09jBNhEwg0EUQQvtjECBtHBODmc0s8wcYypCFXUAo0xGOPmGQigMAlCNyxERwNfCLHmwHEyYwyW9qJExq7MMBBn36+g8AOGbXhTHOHdrgEInps8KSjBLWRAQcMUIoEAkG4M+pA/qRk0oybHXKDyAwbkLpRBk2kigASwAJbzUG3FNqpm+iwocqrupaKRhAy+BnsGkGwVRAYE+hDqx9soKNHo81SRcwlQQQbLDVUGUurHEcsUUG4NZXz569IUEABEmuwQFACzThzbLJgwGtTKQi8U+/BFJiSnLpuorYEuAJTdYnB9RJxbyWKVeDMBMe64Mw3EdtUDRJB3EuBxRQcIpA2za3bLsQhlxrEGvWebK8IAhVKawD67KNNzDZJgcAa9hJhcSX/wLHxseHo8S7QG/WDLxIWG01EKlpkkonLy0Bt08wnW00EDmCE0wCtVLiyj9c14QOs2ERIoY+hb27BBshsazTH/r1wj+HMJzu7MEOJeWd0CN9GAwDAEA20+eYTS1xUOEapEEF14kSs0cCUHSKiAsyFs2C51YpTsPmbfiDCDbOT/9NG0UQorvjpOabOD+uTt1F17LJvXnvqn7d+EAv2AsA7AKYjoqMfT6ggufAE7VC17ADkQ8URtevgih6EQz/QIRZTD0AlqihPXx8KJOA9QaaQLnsBZS9B3wxOrL2+QI8kTn0cjOizB3Tm8IE57ieQEpxMfJyAwwyAQR8dCPBp3utH+MSHqQY8oTWSUYYCrnA/KehPdvkQyBXQIT/JBPB23rvFGj6oOJz9AxN7UAYGRwCBTeBNeHNwn+xU5jo9XPAu/iNYhiDMAbqQsaAELKxee77RB2aM4IllEMQmpiE8ahABBOIDgAAIwoIl+OA1I/CEIO4BQbatImxZtNZArtDEynAFC1zwRBGbdQt1THB/BrnFB1iBF65YYBOowF2z2jCG6VGPCGokyDcUoICuOEADFtBAErwmgqJlUWEHaYMVhqiXoXjCAq9IgiCR0wYpwO54i8NdBT4wxq/kIgqvQMUcMVRIS4qvHBtBwSa4sBejaCAKk0hGCsJVDnXULImcsAktgsGDy7TgFbk5B1QcdYtS1ItoOgSAC2vSBixgIRiYMco5bEAIePQCErOkSjUMRrOwobIAoyTILUjAA3CGJRrw/mAHBoBgiSjg0ib98EW51mDAUxoPAK5qSxu2kYZ2+IUs0VACO9BAg26k4xrVCAQUAtSGW+wgENVAgxj8VIIStNOdvINnadqQhGS0wxt/GQtncCEKtNAAFuYZ05jO9I6enhSlxhNBPDWyCCMQwhFjCQxnPAOLDohDHKHogT0YwIB6iEEGlqLXMd1JgWHpCgppmAQhetGb3wQHQieYVISkRS97Fa1okvjnICFhBEVgABedOauSxuSnvgLLZG69lynSORUWFAIIZ6EBcNAqIb5qtWZ/RYIvEhmyW1QDHrhQg2HMs1dK0etX9AqCKSgLtTYkIgo0YAJx7GEPYmSDA5fgJMCZzgQsGUjhAUONmTwAEQU0uCEe45gUB8RADHxQwxpQ0FVAAAAh+QQJCQD/ACwAAAAAQABAAIcEBQgEhOqMiowMRGg8rvRMVFwPY6REgpwEIz4knO+EzuRsr8YkQ1QEZLFEm8wkJiTIztQkhswEMlRkx/cETIYqerIEEhpku+EQdMRkm6nc7flEbXk0YnwMkPZ8hIwkUmwUIywki9ZZwPwppvwENFysxtwEVpQYGBx0vdJUsur///8UhNiUqrwUUnxMouwEdM9sbmwscpwEGiTE4vx00/wcRFs0mtyE3Pc0ksw8TlwcMjn09fU0U1wUV4cEKkQkSlxMeIccLDMUa680grR0yuhpp7x0uvRUuOwEetQEDBMEivVEZmx0tsxUpswWPlxUkagcO06MmqQKSnocbqREhqQEFSZkwew0boQ0ldcEWKCsqqw0eqTc3txMtPD///8UUoQ0fqwcjuc8p+r///8citlEpuT///80ovFUfoQMiuxUhJIpovpsr+YEbcFcmqwkOD/U1NgEUpQofcEcdLTr9fv///+s2vz///+M4/w0XXAEK07///8MDg9UlrCkpKgMg+GsusyQvNz///+80OD///88PkT////AwMQ0MjT///8kWITE2uwkluyk0PCUttTU5PSUlpyUprSUy/c8doz6/PwcSnDs7uz///9Usvx0tuxUirRUpux8enz///////85irxUosS83vxkorhEtvxQWmQkisxswuT///94ipxMlrzM5vw4ouT///98zud8wvwEXqff5uwkPkR8lqyUwuz///+Myvz///8kTGf///80hsRckrz///8EPG1YfJz///8UHiTc2twUesRcxvxsdnz///9stvREsfcMJDUsne8sRlYshcZsyPRsutk8ZXQUkOwsU2IcJjEMV4z///8cUXdUoODM3ux80/R8xtx8vPRct+R8stxsvvw8e5z///8sOkw8XmzM1twcluw0cp88ovQMUogMXpxUUlT///8MNVEMToQMFBnk7vQsjtcchNQMdswMGyM8nNiM3PY8kccMK0EsS1QcbKw8hrAMe89cp8ZckqSUnagkcKI8bIA8ktzk4uT+/v4I/gD/CRxIsKDBgetKFNEkrp6iHi2mKeLQSxYgWAczaty4kdKiIhX2bTmgRg2QSfw48GvWLM+zGk5uoRpEiaPNmwLpBJIDRpcbKjEM9IhDoagUXkh5Dajxgx4DdPSiWMJJtSAdbXLkFflU4dWrLGDjEKXAi2xSEiT0oIMSSwc6VFOr3qQ0KwKorUIaNPgadmxRXkd5oUWrp/C8N0Hm+akpV6OGfg6Y4GjTRu9evib8lk0qOK0EPQhCzwtyLAewxgclhWAGyh3lynq9GphyhRuQ2xs2tPxQQ/DnwqFDHwNxzA/qgZQywUOB40Ub57CFgHmS70AMReXIpaMgRUqlSs94/rxEBzy4DBkg3g1jLJeSizIoVryYT7lBsD5u7E2pp8vRIFh0UELJDlwMwoIalTgBBT2xoGMeAlXI8M47pOzQXhnZMPPHfNBhAIooYMxRRDU4gYOKE+jE8sY8CJwnQxVVTFiAhTi5l80FSCDB4QvyLPDJHLPQWNUOkThxGIsyQAijOhYUwB5HxRxhSgA56ngPPvgEo82Tje0ggB5BBKEkjFUwOcxNjYhBBJVUIrECCg4o88hxBoGTzDGkkZlEEuokYdxGGiRgTRhtIkEGcw4ISSdBOyxxDDQvwrgnn6dlRMk4rZwRwKZIpHENO2xwuShynCAAwosWWDBpIaIKJEkX/lZwGsAf1+DAxqgcDfPOL1WoquqekBxEBzI3dLAplRe4QE2ruCJHyq6p7glAEifERRA2ExhzbABlMMMOs80OtAMUv/Q5KQAAcGJVAjcocWwaCqygSrg2wRFjFecCoI61/9QiTBfbXlAGNvTe5IEF5iaBLgACILdGPB24GwAjrZABbsHiQpOwwgA8wNgMxggTgMRWiCEJxjfpg7C0Cx8i0DY0rKGEu84oEMbFKAu0Q7kKcwwADAKdgcfM7nZxhBE53zRMEnz0jO4D/6xDwAREB0BEAvMmzREEfE67MABc2CHCKEQ7E08YWt8ETdNfA+CHKzETfYYVKaRtEyl9eo3u/jBd4BHxzEd0MYvdHOnD8sLmjDP0zAFMcEYohG90yOHoIjLC4u5aw4gGkWvERb7onrAGDUQrcYMzinZOkCWUA8DHCKQTHU8aONtNCejowl767LWnfbvPC49edTyoq34Q65S/3i7j1oTDufEFfc5H2ycoXvoyj0Nf0OTAA4AIJn4TnY0xJ2s/ECS4mwO3zDMTcAEm5g9ECttfDyM2AUSHM2j8Aj1g7tda0AABhFE1zWVNexCwAP0WxoV/rAFzAThCCpBmPk4kQVVfg9o/YJYAxiVAAbTT3g5A8L+FAe0fIBMB46y2ivIZDxIr0xsAXPYPSjxshWeoWOrstoMgJGxh/tBgj78IMLJNEQEeBFPdwZjkNIYRhA7s2hYj4nVAu8FBQj/UF7/+gS0iHisb2WDHDlG2A2+844cKU9e6brCtP5gCHsvyHSlkwKtf7WlfqTHasZAQAgWAKm3DMBUTzxWsg1zKGgmQFTuuEYFQoYwSgXzAO3qFrz2xyjEUSwObXmADRlJjjHTawQZ+kR5J/UodldJImqaEBCq9AAcokIcyqrgoOCQDBKSJlKrw9ScopSAbAXhBjl4QASaAAgNbWtQOUOGD0ZyKTBaoAh/ONBcXZCMFVZqPO0CxAGXMIRC9KwiRjPQGHRwDQrpUh5OoYiNsCpM+yiiCG+QghBGV6ETo/nhDLMSETjKpE5QacU+3dEQfynxCFG4AgxD44x8ACWgHsDBQL25BAicwIBk6AE2L0HkMdVQINcmxAQraUVDYVOAJGVADN/bRg2gYBTADqMQtwkMPKJQnNBECQRXWs6hGrMYB0IHNXsqxD26QRA1oAEJuNvANl9QAHWnRgw/MMw/i9HJRj4DMAiJQn9jwRSzkIMcXyCGFAQygMyT4jUYRQBrT0IsudhGFHCyzF7Bkpijc4Uxn1OoDHehgHvoIZ1WuEoxU4GMIQsDMXY0SGKSgRQISSJFbUAHQZu1gJ0N4Qh/AsI9yiAWvm0mKE5zwg6dEZYta8whI9iEOKqjhAJO4IgIHZsuBPLjkBzFBBQQEi7GELEQcihDKF1rQA4pYBCOjCggAIfkECQkA/wAsAAAAAEAAQACHBAkO////BEiAjIqM////TFBUTIac////CSc7jMr8RKDXJCYoK0hR////xMbMJJ71bLTJ////ZMf5KonK0eb5ZqCyJHe3BBYjBFiabImc////DDNLB3bLNFdfbLXsKI3YNGqEFBka////bHB0SJLI////pMrsBDxpFCcvzNrk7vDxBHnXFFmMF0lqBBktKFh+ZMDvJJfqKXapJDhD////dNL8////UpSyNJ/nrKqs3N3gZKrESHqMNzc5FIffUKzc////dMnpHDM6VKfsBw8UVLTnhNjyBG/EdMDXOJPK4efpBF2mZJmnKmiUcXuEFCw/FHjCd7Pm////NFp8NJjc////R2Z3rNb4////BCA3JEFOPKL3XLrkVIeXWZ+8JE9nfLLcNIa8dKbUVG6ETJbAHC40////BX7b4O/8HGaXBC5MDB4kFCAnHCgx+vz8F094////HJDk////////FHC0JG6kfHt/lJuf////////OH60DBgdMmByPJLcSq7upKq0rLrMvMrUNHigVH6M////VF5ofL/4v8LEhNHmxNLgpMLcpNL8////LH60PHCIlLXOGWCZPEJMbKLQbKrcJDI4XJanvNrxXI60////pKKk////VJjQFILUNHGYlM78fIqXj5OUVFdZVI/BXKvOBHLN1ODw1NLU////ZLrn////GX7MWKbMF26s////HDpKBDZgDF6cKJLiDEJo8vb5////PKbp////TIKcit75////cYKR////dLrMPH6c////RIqwKEhk////////////////////XJ7U////BIHm////LGKUXIKk////DI70LI7Mb3Z8ebrpOGKArLK8fIKMlKKspLLEvNLkbLr5HDNFDEh4lMjwTKDQLJ/2bMjwDFmUFDNFPGZ0TKr0rMbcDDxf1NfcHFaEbMLnLJbkLDhIfNTy////bKu8HIfZPE5cfMjhXKbkXLLcDG64HHe6PJrYDCEuLEJMLFFhPIi0fKXHXGp0HCAkLHCkPHaUzNHUrM7s/v7+CP4A/wkcSLCgwYFKAGUYNAXYuA0bvrlqZ+WTNCUHM2rcuNFNogzKWnyx16EdAwZaZrhSSekJijUznDhww7GmTYEq8IEoBw7ciy+yXr1SQ5RoFgRZjqJAwQbmABU3oxbMCQnEGD5vZJ3YKnRo0SxqkmZxMXYNmz1rpkGVatPNozT8eCgTQFcrV6FFw4YV68LFhb8h9rC5Q5OtRiV6OnVpQrexgK131XgFK3Zs379/iRCJRM7wQUWsuvTyhgGDANOyHgsQ+aWdPZP1tLj6toHvZSIXNBPZk8nzQDdiZFSqU/q04xc9pwAdh/cVxCfYZpR5YjmzbiIAnhVmO2uTnhuQlv6UNo2BhSMeWMstoxZIiQo3blToCERtjCsEKNqgyII5N3bsoczCHQk33LDEgeJh4M15L7DARAo39aPLE/QwZd1/AABQgIA3ubEJGWQsEQ+CGOyzWBOPcCjVLNRgY5Zf/mWYYQHbcTQJCV6MOOKBBtwCCT4qejbLJy6cdZ2MAIxgkwlhjBLPjkuwckMvdUDoW0HkoLPGGrohCUBvG1EwAQR0HGHmElBUso8oQV450CyFXLBHlzIS0VlGbsyzwwRnHkFHBTJIUqObBLnxjJyaednDoANpM88qZpoZTwUWSEIoR4fuAQB2SIJykAofIMEBKZF6cc8mjF5aaJyaehnCWv4ERQHPPKSMesQ9XjTSpqoGzYIObl5mV9Asoa4wKilQQAAFBbzW1E+MXsL6jyE/zLPCCqSQskoSYDRr0zSJejnAb1S4w8m1HLADgTy7ettrG+HKuEBhFOBQxLXYjuKMNu7adMduwR4ikAfmxHLGtZzwokqq/Q7LBqdIKvnPOYiccfAKCmwTRcM2OQEwkgv8gwYOXJyBzLWoTMAsxxz1E6+MOlzhhx/IWOwDEuywbBO8EGeYiSHmxGDyGVTA847ONeXzcnZDGIHM02f8oAC/SG/078sF4OD008hwQYUlVW90yMcy9sBNOlwjg0QcaIStkQ7QZhgCN0Eg08zT7vjQrv7b/6iwW89EPNBN2oj4wLDbsyDqpeB3P40IMoeH7QaiPT9Qg92O6833QbPsdsGmGe5x9tN3p8P25gbBPWeiROizBS6NI3MNDmCjTtAhF4QQIxHohAN741zUQrXtAt3hApfXhWJIDQ9gXgsM8BA/UD5b4paoE4sEf3czzSASh/QCYXOWf5rlgEYtEnD/tBGxrGx7P0W6YL1mpfzDDS6YN1OEH9BIrws9bIARbighkGswr3EPSAc79qazWbiEHpjBzTMEQgE/FEF9yAgCDobHN2ogoA38wcweHCAQNzzgd3bDgQIZ2LBZuEI/l5FTG7aTAAnUYnsZrIUhNvcJBAjBMv5+cYGn3vQAI+CQG4VzH9LIoYaWAJENbJDWtGwYOxhwIRYs5NUsgIGNHyalL2vQRa+KyL3tuaMWQ4jcpdwwhg3MAAFIGQsKniBFgXiiCFyIXQwQcQ4P6CwDG6jHN+KYBXrQgxp4qsUBMYcDRMTCA2r0jBsAWY/7EPIJ7TgcGsgYuy244xxDyCJ3BrEBYGjhKwh4whPupJFFKBJzT8OBOxTwASW6iRwvGEc7gPEVNTwBAX+wyTW4IIG0ISMG7hgFJ6Ahyo3MAh8nIIkWvBKWbzxBjG0JBwxgAMszcAIVSKCCKrTRzGFVoxwt6EAHgpIXiVghkr/RJje5djAqIAECzv7gRBSsVJMUSOINb+CDFb5wgq4QZRwzeCd3LNiNxllsBWdQAC8gkAQoqMIY2rCGEmYBn1kowRqPuAQkvPECRziCD4+5yytkM4ZyFsQN16iF6aB2sRU4Ywe8WAcJwiCPJyEIFmmoQyd4MAj0CEAWdinoCRySAXi28gHuoNlDr5UtDoSBBKtYxzoqUIFKMKESl7gFP5rgDcek9ATpHIc0VEWBLfjhkzXFVraO4FM6sOKusBiPcRpzAgFkowPAiIYOmuUGbcSAC+Y4B77mKqkjiIhEejUNXQL6gjc8wqkrgkYsioAEBXAiW6RyrGgRlCDJYqCkWMFHHb01C0P0gQoQQE6FAibAgUgdyLEHKg0khsoPZSgDSG5zgyU84IwJkGEUO6jANsjwi+b+oguXMAA/gsqERGDWXWjwRxQ24YxGNAIK8qiDBcLwCzEoAiOECggAIfkECQkA/wAsAAAAAEAAQACHBBEdDIrkBE6EjI6UTE5UTIqkBG/DBC5MTLLsJE5kJDE4lMr0J2+cTJ7KDD9hbLTOBB8zRHGBzNTYJF6ALI7MBH/h////B1+jbMbsb3V8JIDAM5rfjLLUBDBX3+v4bI6sVKHic8DZFCAmG2+wRIKcBEBxNFBe////rK+3////PGBqLI3ZBBckHGCQFC9BJHa0XKvQ////BHLMbLTszODwG4DJ9PX4OUFJYrTZBCI8NH6olL/hLFdqbKu7FzlUXG58FDhMWpisxMzUHEBXPJzaFygrRIakFEhxKUhTFBocXKK8PIy4DGak////BjdZbpm6////////////////XLrqHE5vJGaQfMje6/P5TK3iGni6NGqIJJfqjKK0NHGRRHqU////BHjWGobXPKTnXJCgNFhkBFicOHakfLToBChH////////VGFsbLrWJIbMpLnMfI2cfMHUJEJXvMXOOmd5DBgfGWehHDA8////pMLcJll3////wdvvRJK8////DBIVpKSkIo3arMzoRGp0VHqcfJ60XKTgjJqskrjZBEh/vNbs////lKe33NrcVIaUXJ7UfIGESXiJfc/rXJbEXK7c+vz8JDhALGKU7O7vPJ7sVIOpfJasfL70fKvUVJKkdKbM////TKbcJHq0////ZJakHEZgfKbIDHbETKbUDCY01NrkTIacbLr01Ob0NEZUZKa83Ob0xOL8JJLkqNL0VGp8////JlFvBGGt////VJbUFIDYXLXo////JElh////////FHnHLJXnUFZcDDZMNJLMCIbkb3qErLbEfIaP////PF58TKr0DHG8dLXEDB8oLF94dKLEdMfkLILIlK7EdIqcTICWLHWqZK7MdLPn1N7pPICkHDdFRJvPTISwHElsZJ60////ZL3k////lKCpPHKMTHefDHfPRKbcPFpkDFqUDCc8xMbIXHiQDEh0fLrpPGqUdLrMLIbClJqcLHqw5ObnDE+BVI6iDDBKVLTqLE9fVKDGFD9g1NbU////DH/c/v7+CP4A/wkcSLCgwYH03sChZcKHC3WpXNxxxQYZClgHM2rcuLGSEGoJ7m3bZqnkHRdFUqZ05oyFJWPrKnGcSVMgli629iFBIgfIvQNpcggVCqEoC2ciRNSpo2AApppQC2LZ5K2XCSROOgAFmqYr0aIQWEAAwIJFHWd1REB6GpVmpWlVbKEr5SRrh7tpuAb9GjZsWQBk6/wRIU9mW42w3tlSkaBE3bp38Xb1ylds2b+A6wC40ejwwTf16KgoQdpx1tPDhgyxJKekpZPq1PW9DLj2nzqAPA+s9ERPhCqlHQsvle+qD58HuMZGeaeIUdq2AWQw3LZSNysR2iVKFLyE3F5OvP6xYzQHlo1KlWw0mjPux7YcqYqkqgMd8B8Awmy0tdGNTgQB23FXQjs8qNDLER+oUpMEkLggQioiYFYbAH8QoF9N1pFDjgAABqgHHbZUMc2FUdlwiAIQiNDShPYRQB1HT5zxhRkcdkgHHUcU8uJhNiCTYoQsApYBTXkwUA2NNdYTyTNWZKObQfzIkRYLLN4nD0et2GFPOmZ0KUA6kbSgCYlPEmQDGywAOeEffyiYUSXa2MNAlzSmQ0gLT+xY5m7GpBnkHzfoKRAi2hhBZ5eOWPHEnhxl4GeQAxyExQtB3HLoF150Iyiju7GRVpVJsEVQJ33ocMupZlhTjTVkcnqQDf5SasYiJFJpQcqpljJBhh1uurqRBFOu+Yeo/yBSKq63rGLNJ77ShMyjE0YqUCUa9GAAriMEYcemzZp5h5r2KWAYH/H0ccu1txQwzw7d0iQPtLWtI1An+mhwrQHMfDMCt+0OZIODVNo25D/xPHDvLUsYwWy/MxmD1JoK/OMBBfoYcG8QWtDA8EwSsFBEwLXRIwg38ViMbw9abEzTNiL4FRgg2MBg77XxTKKPyjP9oCJmdUACAjzmmNwHMezivJE8EHx8WR1sEAOPDDJYrI80ihi90RzOKF1WHTes0AbUUV+jhQdWa9QIBHdcdlQRgeAgQxhRPwAMFmVnZEPSzyHV9v7bUMNzCr9GV4L2bBDU0TbccPsNOM6C35GDXxA4E0g4YVQuAzzAtFq3TTnc4YxRIqgTTAiVw91GDWRvXlAjOVgCVop3bBBH6WHAQAEfqhc0RxquPw5BKq6AcIU/FVSOChFF5y4QI/f0XlQRbLgTTiClbwCDIcoPxI5EX6mDzCxZjFFB8WLEIUb2Aq3mwl6xoYBFMPgUX3wIYrSSvQQHIEGZC6nQ808wkhhfGCqAClRgI3vUAIIlJtM5OQhkBhgIhAADQb/FtcsG++iJXu5gDIHEYgzxE2AbNrCA3DFiGK5ITlfu4QIhTAuA4xvfBtogBgv6ygbeQEIpVMg7OVBnAf74EJ8AQ7ABd2xuE/swgV3SIJJCEAQLXAhgDGVxBV3Yz2qqcAIP5HCXA3RgCD4g1j9YEQ4hjg8HlFiBDcuEBT0koAyQ6cAw5EANg0BREsWIYQWiQQQQrNEzlSCEN1TggNN0AAn7EKNAFoCAEI4vEFcIxgxU9oEjqMAbTihBB5ywjyEw4k2ZkAQX8ijDKwRiBn/E0AfaQQcekKYuJciHMrjlgSiScnxEuMIK/NgsG2iiCnR4xisdY4tSOGkjsxgDBm4Zhg1cgQgruCKjaHCJCURAmKVxQjGnQRNWIIAKeqwAJCnhD3doLiqVgEY9rBAJPWynNLawxQcwtIxwgPOW/v7AATyIIYYFnJMjNkBEC1oQiUjUI0CkqYIeyrHGSiyDCuDUYzPj8IAV6AIbGqMJDaBhhxZ8wRFbqJGHlPGFf2rEofjAQDgHeA54PIAbuqiBIXbABw9gAT028IAiEDGJETCBAasgAzm4xKEA0WECDAUkK8YgCVnEkHbEuEYcetAAYmjhFMy4FxNGMA8j2IMU9vDCBbyEpHqQYwIfSGVBFhCMaGRBoqUzBwUacI0HNKMHeMXrN76xhHnY4VA0opEVvtCCN3DKA2NYRjQ2ID/awc1ip9CCZLXAhGtdAFdjpROYGPCOY3KqEgvgAhUwEIwB0o5vBoiaxZB1qrFayg5f+HmCFTig1plgwR3wi8Y5xFA6sKn2Xgc7FROsYY8zMAAadNsYFnYQDCKEow2ooIA/wGay1RqACVrQRhAKYA1rcCC5gYvFDIgRDFTgoA04gIE+GsBefShBCfpYwgvm8Qk+1Paws8AGCIixAjfUoAYakEY8ctEJQaSuTAEBACH5BAkJAP8ALAAAAABAAEAAhwQRHRSK5ARQjoyOlExOVFSMoQ9wukyv5wQvUitQWSRuoJTK9CQxOlCgzGzE5v///26vwgQeM////wRZnv///zyKtCpfekhwfCR0uHCgxN/s94yy1P///2yKnGyz6gREeBQ8VBQgJf///0yi1ARirDic26yvtzlfaUR+nBRajBQuPCRASwQWJFCavFy66////zR8pBqByySFxyiX6QQiPBpknszg8PT1+BROeHTK7JS/4f///xpDYFSWtFyw2kyp3Dk/SXK70f///8TM1BcwRBQaHBSG3CeN2Rx3uDxvhwY7YzSV2VxufGyizGyXuAZ2yzym6TRqhP///zSEtv///1yMnP///3zF2uv0+QRmtCpHTySS4IyitCRmlHSCkDxndEqEriRWdFym5BxwrDJXYwwnNBQoMUSdz////wQoRv///////////1RhbP///wk2UXyu1ER2lKSyxHyQoHyw4P///wwYIBRWiaTC3Hy77v///1y27MDb7xQ+XAwSFVS27LzGzqSkpKzM6P///0RqdGypulR6nDRxlIyarIy23LzW7JSnt9za3Hx+hFSDlHyozEp3iFyWxPr8/CQ4Qf///yxilHygvOzu7lyg2GSapHyWrFSCqDRabESUwGy2z////xRilBxqof///3y+0Cx2pNTa5ARepiR6sHR+jBxCVAxqrEyGnBw2PDyCpNTm9BxWfGS22TRGVHSq2Nzm9ARqvMTi/KjS9FRqfDx2pP///zee5////1SW1P///2y69BRGbgQ+bf///1yivyRIXlBWXAyG5Bs2SDSKzKy2xHy2xHy27BRqtAxQiFSx5AwxSixwmHTF4gwfKyx1rJSuxHSMpHSz6QxFcRw6TFSh4gxlpxwwN////2S74f///yyFxNTe6RxQdnzN5mSx0VSp1jyW1ESi3P///5SgpyxojHyEjyxZcRwqMAwpPkx4oFx4kJS41lySpKS6zEx+jBxaglyWqByG1DxqlCyS3cTGyJSanHS2xAxenCx6vOTm5wxqvP7+/gj+AP8JHEiwoMGB/eSku6WFiJkIdqKtA9IGlYl+BzNq3LhR0hAvK8qYMbOuZIiTIaLZKWKHBQAGqPBJ4kizpkAsiIap0DZJRZloESKwEMqiaFEALPwofTngks2nBbHMMWZsBRFnNLLSCBp0qFGXLgGI9QOgSCOnUGtKWpQqVcM0cLVm5SqU6NGwYgGQDZFvZlqNjL64vYYAAdzDabTS9fo1r2OyQBj9PSjnV4JhhTMjTsNOBRFtn0uyMxMtmlHHj/0EmjxQErVfZPogUJI5s7FrK4wRYcdObmkaIcykPIo6b6Mbk28YCkNGifPabwY748FkEaBZlyRJuvQN3zkmxiL+RDuJtzgx5FAlxVFHBpjz90qGJbjGg9oQm6VQPRxeXC8B9Gq1ow4nHzgHzGw8kMHDMNMA+NQN52hjRwh29AcAAX7R5IQFX3wAzIcfqpNAH5o4+NcN6UBURH9+MFFTPPMQ8sGMHlpzAjgWlMKaQUOsMGFxZK220TfzQMIMjR9Y8wU4hpi4o0A3tMFCCED6oWNGkuASxzwCfNClNRfgQE2GTxYkCSpTAgkEmQS9g04SXcZ5AThOlMkRmlQWl89BWHThiACAdpmEBWCwaaeZbfz4WAhoEWQJLl0EKoAFSaDj5KEF3bBCSqj5gUpUoRQgwASA3iHPHVdiutEQ40XgmB3+djT6zwa4PEMqqXF0kYGqNaUz3mkA2DFAa9LQY8oEpIIiTw2X8pqpQy0dZQcDfvFBSivIIouCAvE4W9M5EZjxVUv4CNREAWMcO8E+VYBiqLcG3aDCOnaxEM2n//CTCQmm9AtDHE3AWxMq0TxklB2T/KPBKRWQwO8E6KYqsEZD0ECvV6b1I0gFUzhsSjb0jDFxTcasA9RQEJkgSwtIOEzCKWBEMjJNTJihQldTesFLIapk4TPH78zM0Tmd0WVGG8cUQovPJPRwiiJCbwQIDQxwxYIZsXijNC1LC7OMBlEDlsYkWwVlhjExCMP10oUYgEXYGd0w9lw0mMFO2mvTwvP+u3BLMnYZc6lAgwzjPEGL4RAYwHfYcq8QFw2CE34414m/DXe8aazATmKQq3CEPk+ETos4SIB9eUGzpKHFYTQQscISQRhuuBgy8HF6QYA4s/phxiSATeyhP9FJJzrcThAXIKxu2Ni3KCNOPcEfMw4mxg/kTiorLI/ANXPYUs4SwccQRAzVC8TDiJk5Y4wcGtzTTPBPeBCDDdWXogQZzmQGgjMY3XMF/CNogAeqR43zzaYww9CCQDzAjQAELwDjW9zEbgAbHtAGAW8YRgcEwgddjKMYwfPEEop3umn84gTPUUIf+nCff0hiCeF4AgifMANP1KNZApMEDjiRgPcgIAH+4MhQHvagCxmGzgHkUMblLIGDE1jDPUoARgK4QBAszAAaIAThFoJgBFeErRQf+II6PuScMPBAVv/wBSx0kcUn+MAHR8Chqm7QBQ556EPWSIAmDGLFHLSxGHkoATYkaCdJbOIVYfKQhzgBDjQKZAEHOEAbAxCOIwxwYk64AyQsgCRwhGEaWNJFDrZQjCwu4QqWJORkJJFJSETBSzNixhfssTgNbMGPpQwdFFA5SF7doAApkAec4hTGO0hmI7aAgi+yCEJdhIMc9/DioWyggC5s4hCS+kAUXtEtmvjiAC7IpRav4INiKEOVG7nBI/aBiyo8I5voQEed1HKAPfyhlOL+hMUVllCPBcgRS+8YQ6gKAIpbAQodh0ABOgUiCWy44J4yNCU0gnAPI1SDfjWxgSyWMYZV0ANb2QLUM5IAhn8aRBI/aMYymSnDclwhCGcwQgzEoAM+aAAL2rmBBhShA0wgQRVT6EEmKrCMCagLWXE4RKFW6QsojFKcoSvGEvQZhBHwM3iHMwASvFEBYRSCHhXIBr/6dax9rEIBGVhoRmwxgxxAAZ+5jOoSxgGLUdh1FPrQRxD0UYgWTAEJTHNZv0iBLhKWSQNQOEAOZsDMGc4wdPWIgWQNl7elBXYMPYBBK2ahKkksYAa+cABjowq/0sqOsnnLAhJ6UAEMvEOtT8Fugi9msAe3BgCfRnwsVkVHCwMcQxgVOMUjLCcwLHxWmdCAwmhNe1oZkEMYwpiCNxJhUkxJohbVmIEuDuALaDhgD80Qww+w4QNYeEAYZ/CGNzzAB9jySgO2yMMPdHGPLdRjC0c4wjGwUQ1BmO5JAQEAIfkECQkA/wAsAAAAAEAAQACHBAkO////jIqMBEiA////TFBQDGisRIuvBCdEhM7lK0hRRJ/XJCYsBGi4xMbMKJ72arLIZMj8BzZb////0eb5JHe3CxcbBFCPLFdsZKCybIacRG55BHbMbLPpJIzaFBkaFCczcXR3NGeCUJSy////BDllpMrsRKbkBHnX7vDyJlh9zNrkJJjqHFJsdNL4BB40BFyiZMHz////FJL0JDdE////THyMLHWprKqs////NJ7mFDZJ3N3g////FEhsdMHZ////Z6vC////////hNPq////BxAU4efpCx8pHGCSFD9ZWbPgVK7cN5PMOGFxdLTs////VKfsODg8////JFBpGHK0dLbMbJKsVGyA////rNb4BX7cPGKEOZnZVJrBdsns////O4e0fKXHOJLYBj9r+vz8KGCA////4O/8V730lJyjVF5oVJ68////OH60////HHe3////////XJWkBGGu////////////bLrWKZLfpKq0er/5PHSP////GU5x////rLrMvMjQVIugVIKMfJKsv8LEJDE6////pMLc////////bKPPdK7clMjwXGZsKH60////lLXOvNLkFFaMfIaRfHp8vNvxNIrHpKKk////VFdXPFhgFEJkj5OUPEJEbKrc1ODwNDI01NLUZLrkFFqERJrM////hNr3////VGZ8VKbMGX7ICE6EDC5ETKbY////////b3qETK7s8vb4////DGKkPKbs////fM7kerrqdLrMaJq0JG6kPI68////jMr8KEhk////GVWCTJPPFInl////FGKk////////BEF1LGKQTGJ8////////DI70LI7MNqb2////rLK8lKKsXJ7UXJqspLLEPHqcXIKklM78fI6cbLr4VJjQDEd0FGehDCg5TKDMbMfuNFpsTGqELI7cHCkxPGd3rMbcTKrv1Nfc////fNX0DFyYbMLoHJPoLDhINHOcPE5c////HDdFHElkfMDUjNXmHDE6HD5RXKvPfLLkXKrkLFBgfLLcXGp0/v7+CP4A/wkcSLCgwYFHcMTSJAXJBwAAjHyQoikWjiMHM2rcuLGMgxCGjIgcCbGkSQAMQhQqw7GlS4EpOhmyQJPkyZsmGQhI8bJnwRSUkFj4YGGk0Yg4cRqpxNOnyzJqQCBhQLNm0aNGkCY1+aETS6caecRDUk7oCwsvjKQ1SjLr1pNSeIA9qOfbPRAv8qJVy1dv1ao23wLANHdgGUoI7iHJyzhtXiQO96a9OrKoYAAhvjqdhaWVIQSNGYOQisQQFjWFeKSYVSYFqEJqHJU7urXArM3L5s1DwPsF6Bcgyn1rVUnUy0CxZousfftlGSzz8PEGzbsVDRD4pjX3mUJNuatJC/5o5qgBH75W03nrbqVt+9xZlaoqDeHSmhIFrSQgyN8K3zxgKxRmkCg0GEFZSVkRthEoSsQjgX4QKrDDOO4JONAsaxiBBG0GypVRGVzww8mDJEqowXgWElRGLAayJZInKA4UCTBUkPggMEpokCJHLG7YlhFqHJSCH06QUcKRElABDBYx7qhiKhYIZVSUTRG0izh+lGBkCfTwo0KFTho0yztmTVnJT8Isk0wyZJCRjBOcpBNmSyu88IFjBlJJ0DRmtDDAmmSI0wIhc7oUlJSOWdCJYcrw0c2fybBijjBgFirmXXnxVdpXKyQhwgCgDuCECpFY6pIaZGV6FhIOCLTLBv6khMrKBpNUampBZYzWGFqxCMTLIBcEy4oIIuxyq0saBNcYCDT8c4Qy8AR7wQB8kBLgsRyJ8sI9jKGFxBGIwMOLtJMMkgS2LuFzzzeMfQOCNGIckAQMwXpzzRzotgRdK8tqMIIg7FwAAwzwwFNqvhtNY11viaXixhwD03sNL5IgvFEgrbwz3QvzxFMBxHQMLIg3GFmcEQ/9pRddBSPAEDIMc9Riq8kpSIDfdDvsUMUIdPRMxy61NGnyP2W0ooAE+SFwHhwjNOBzBgYIbfIsEvCjHwISzCNBBWw04DUdUM9scc1WP6i0Eo8E4bXTGVRR8tAFHVHCJjYCw8klEKzdAP4bj1QMd0GBcCKOjQqIw03eHHhdyiUH/y2jD4OX8CA/WOiTzyocJH6JF9Q4ThA2VGBw5JEYXGHC4olzsAoEq3g+kDBO+OFmCZxQAQgaz6iSeQMcsE6B6yt0s0E3biZDhQ8YkaNL5omDUwojrl+hgjlrrjmqQB104EHmKKyiyypiH1vGJBuYAWoyol4hkCVdLIACB+930EQjjkcyiQ2hDqBCC36XkUc9W0CBAJ+BB/DBbRZJ4IMIWBGqDShDM41ggg4EKEA8dCEXcBNDEgYxiQswkBQiEANBUuCOHwQwgP8jxu8sBopyjUtafEhClQSSiyXoYAs4RMES8kGO8DlpFv68uIYgpHUBUvBhEWJiwRdwyMQvLCAKUnNSGYbBi5HRS2LKmOFAGiGLJTBxC+5IQB46gK5FeGMO8IgYDG5gsA/pwAUscAYTdYCLMUaxMGUw4y4OoEZ2CMINUkMDC3BBjC3IcQu2EGMUfOiTWXCjChngY89ERjKOaAEa4TikHHVAhBPkYYU7osAj0nYAn/XsADdAhEu2kYY0GJKJgxwFMfbASI3Moh8cGEEQ3KA3OoQhDEh8Cjpi4EpnyNGY7cCFDtzRiFoSZBaNWEXaglABvTXADb0YwR2fOcwYHJOJ0EjAF3RAjFyAkiMUeMIqPBAEK5QidWvL5jCcKZBZyCINmf405jdlQYQvnMAdxNhHIyyBBtaUYRZosEQj9rGKLTQBAroAB+Z4tztwDGOecynDNqCxjhl885i2CIc9cLEEW7CgkF/0QBdcgYcE4GEBmGNe4laRj0t8Ypsb0cIDTiELfX6UGNBYwheIYA8iGJWoCWjHErqwPQHKtAFjuJwJnIQGaKRhHQ845kdx6AxizMAdeSAGSnNIQad6oAML6MLbdlSGbDwgAi6AhlYN+c0v2jWHATwrEzzQCJxuZg9v5alHt+rTV36RGLZox15zoUVLzeIXD5CFC1yQhqzW1bCGjGw42qGDMdDSZGWwRAygYVUXnMIF4YgAMfHpgnWEQxY6YBvBEyzhV1OhQQt7kAU0HsDb3uogCtvQAhp2FBAAIfkECQkA/wAsAAAAAEAAQACHBAUJBITpjIqMDEJkRLL8////////DGKkBCM+JJrvhMrcBGWyRJrMJERWJCYoZMr8JILIBDJUyM7UbLTQZJyoBEyGJHi4BBIa3O35PGRxBI78CHLERKLcZMLsRG10FCMqJInULFJkBTRcGBgcbIacrMbcKqb8BFaUCxojLHKclKq8dHR4F1J8FILUxOL8RK3sFENfRJK8NE5UdNb8PJXMBHTPVMD8bKLU9PX1d8DUBCpENJvfhNn3JEpcNIO0fLLcHCwyVLTqFG60FDVIFI7pFz5cXJqwdMzuRHqUJJTmFFeHFJb4BAwUIWWUNGKMBDtjBBwxrKqsCWqxJDc+bLPsZKnBBBUmNFNcNHqf3N7cVKvZBHzd////GjxSdNL8BIr1kZ2nCkp3WGFsRIKkFFKENHqs6vT8GIrkecH8////WaLBWHycDCQ1LEVU1NTYBFKUJH7BOGyEVKDgfJasrNr8BVic////Gklv////lN70NIrE////XLrsFJH3////DA4PrLrMkLzcvMrceKbMpKSoPD5EHHKswMHENKj5lMv3NDI0PHOMfIiQxNrspNDwZK7M////lLbUVISSjMr8JFaE1OT0eLrsVIq0lJacHGqgLG6YlKa0+vz87O7uJE5p////RH6YfHx8vN78////LKL0KIrM////zOb8////PKLk////dLrM////JD5GZLro3+bsBF6olMLs////NGV/S6XrZMb5FITgBCxQ////BDxtBGzBXK7sXJK8////THJ8LI7ZFB4idI6kaa7BVGZ03NrchNLobNL8fI6cbK7kVIqcDIXmLJ3xLIbHHCYwzN7sTK3kHEVaTJa8PJzYjNrxLEtUPISsXLXiHDRCfM3pTHmHLJbkLDpMDH7UPKbwzNbcJG2fLFpsvNLkDFaEHJbyfNLvDFKJLH65PIq0TLX0FGWfTJzIDDRQDE6ELHq0DBQZ5O70TKbUbMPpdIqk////HIPPPEpcfNj6DHbMDCxDHGusHI/sDDpcDB0tPH6c5OLk/v7+CP4A/wkcSLCgwYH+oqwQ4wDYHwAA/oxQJGZFFH8HM2rcuJHToRUOIIocSRKigxWHOHFcyVJgJwEhS8qc6UBAp5Y4C3YK5W6mT5/uQt3MyZITphE/k/ocgUklUY3ECjEBMHWq0qRVqRYi9vQgoQtVwzIZS7as2bNZx1J1R6jrQE4rmPxBS3XsXLRl7+Ilu8IpURximIBFe+EP2GZiMB3K0okTp05ZDmESA0SwYbwXLojB8beeO3eZMwsWDPpDKAktJaz4cOHzYLCw/9Tzu5KTGHcoYDOxshuFOyCYOP/FBMSK7wu8eV9AgUIM7Y0rmGe2Qt04c0bC3eJg5I4fCivIqf63trKC5SYUwKpXB4ZiG2q3Bd1sQ/9dPfO2G4l9+FAfinH+m8F3EGAo8FcdClB8wAZXGXFSzwdQROhfglCE8pyAb63Bj4IRWgFFgTJcKBAYbAAhYYRAsLEGhhyFgoCJJ7LxARgHdXINjBKmOIyILL41zAc4ojjEUAQxgg8QCCQJxZENZNdjRjg0AMQQUCSJABv4MKLTEFNYmeQUOjD45Ebe6DAFPl5eMyRBc1yDppVdDKHlmCsdg0+XOiSJzzVzvNVGK7focMst+LTShZN0QtlFK0MIqoMOrTTglARDdDEoArfEuUmi5q3ThqCDDjHEeyQ0sM6gt0TQRhGIcnoQJ/5FNDAEqhE0QIJAnsiQaqpdQCOPqyzJ00UPu95CjSf/vBKrCMyKQM0+3gC7kjfrXNGsCNAU8Qog0EDT7BNX3CEtS3dcAUOzAzQAyBwhDJDLEyLAEMKK43K0hic9PKHvEyHMsQY4A4iQSy6eeKJCvRyp4Ak4Aw8MzhpOZNBwLuCwIAjCGwkSRgZPNAyOE5RInEsYuWQQxisYa5TFAB6QPPAslLBgci4V5NJyqykLZEYYLYdB8iwssDBLBSRX4EsFPKbMCc/s0BwGOEooEUcYFVR9dNIYc2K0ElVX4IQSTcTRdQXZlINzzmaQXU7XcYCdTdVvVICEEijnbNAr5Uiydv7VizSRgiRvBC53ExfbXVA4SgAe+BtIYHGJJHUsnoImkRheUCTfAH5C4JJcMggo6dQhuiag8GI5Qbxg0c8JotcxxiAlYKFJ6wckk8/pA2WSTAqtZ4JFCe/kMwYsxMNiRCbO4O7MARQcUDwW37zzTztGFA9LNf0MgvsNZVRfPC8WCHQDL0LAsgAshhghBNbAciKEEeaYv8AC0dwgUCPtnDP//PAHYnkghqiCFPbnAyw0QiCcMEcVFjBAKcBBDYZgX6I4YQEjxGB/CzACBPwSiBj4QBcgXEAVfPADu/0ADlXYgBRACIcY+G8gZkAhCFkojHtUImWn2EAVYjBDXagBDv5mKAgy1MCMHqpBDRtEGA6YEY1H9BAO6ihhQcwAgQlsoAZXrMEEaCAHCQqIE3JgxiroccUrPgICQTRILBjAgBq4cQP0WAUIkDEuZMSRBm50Iw2kEYsG0UCOeazBL3IwRy8ShRN2XAU8AtmCR9AgaRigRzy4sYUaVJIGOfhFF+mEAzmAIAdasKQltzABEGCAI46ggSu2wEpW7iAH0vjFDXt0il/sQAHwaGUltbADR7CECvCwhi63kARsWMMWljAklNDADS1gQxrDfMYzqNASTtDCGlpoZQC20AJXYGMHZ0iEMt8SC31oQwHxSAIrt7kFDliDFsq0pjX4EIBt1nMLO/4oRjzAaYlTtOQUVDhDAuJRjGfY057P4AMtzqYRawYhHsqop0SV8YxiHOEF+tDHLhIhCgzgwDE4wIAoErELfZzhBUcohjXOIFGJ8iEI8OwKJ6iQCnKMo6US7UY8pmHMbiSACBL9AhGS0I0gHGEaR3gGUHFKhCOkggrjNIgjEoANdLT0CxIlAiL4cNRpePWr0yBHB15w03piFasBeAE2EuBLFmGgG+goBiIC8IW61pWuZ+3DOBKQgHEs4ayAvesXEuAFdHTjlE/iRCISUAsvzNWukLWrBjQQ2com4Ai1WIY4OWUGNCyjFvZAx18jS1nJfqG0de0DAeyRWTSkEVhmmFQEKQgwA3vYABFLKC1lJ2vXPpjABjPwAgGWMYnX1osTLkADKRABXB7YwxgPqEUtHmAMe/BgBrclBRpEEdUxmYEOaAgCIhBhgvKaYLlBQAMdECuggAAAIfkECQkA/wAsAAAAAEAAQACHBAkOBIfvjIqMBEZ8////TFBQR42xBGasBCZDhMvfRJ7UJCYsKkhS////xMbMZ7PQKZ/1////ZMb4KInK0eb5Z6GxJHe3BBYjRG16bIecBFicJFZ0bLPpFCcxBnbLJI3cbHV8JGaMSJLERKTkFBkZkazBpMrsBDll7vDxzNrkKVd/BHjX////dNL8ZMHyF0lqRHuRKXapFJL0////ZKzE////BCpMBBktrKqsJDhCNJ7kGzA43N3gVbPjTJK0dMfhBw8Ub7/cGzhIGYfZh9ToVKzc////4efp////JGeUVKfsVIiYBl6jdLTsdIKR////rNb4F4LUNzc5VLv0CXC6PJbMLIC4VGyATJi+Bj9rPGKENJjcFD5UBEyHfLLcfKXHNFZcPJLc+vz8////KV+ABn/bBCA3NKH3////////4O/8TGJ8FGagDDBEV6HBKVBjPIasFHbEJ5LhFCAmGU5wGY7olJyhHCkxfHp8pKq0er/5VICQZK7M////////////rLrMvMjQPGZxfJKsOHic////v8LE////M3GYpMLclMjw////////PHCIbKPPVFhalLXROWBxJDE8vNLkfIaRWI6wvNrxXGZs////pKKkSazuHGGSPEJEFHK0VJjQFEJkj5OUjMr8VI68XJSk1ODwdK7c1NLURHKCZLrkZ5qx////dLrIGG6q////F37KfL7UbLrY////DB4kcHqE8vb4TIKc////hNr0////errq////PKb0VGZ8////J0lkVLTu////FInn////BGKu////////////BEJ1J2GPBIHmXJ7UGlJ+XIKkNGqELHKsSKbcLKb0DD5cLI7MLH6srLK8lKKspLLElM78fI6cbLr4NFp0DIbkDEZsDCg7bMjyDBgeTGqEDFqcLFdsLI7YdHN3rMbc1NfcfNT0bMLoTHuMHJTpbKy+LDhIPKHffMjeXKzQPE5cXKrkfLLkfH+HXLzuXGp0PJrcHD9SDCAuPJ/xVF5sHGidHHe4HCAkbKrc/v7+CP4A/wkcSLCgwYFHcMx6JGUOEAAAgPST8mgWjiMHM2rcuFGMIXILIIocCaSkSSCSyDkQw7GlS4EoBIQcSZPmyW/fgOQQgOKlz4Io8DysSfThyQslb8z5NodSz58uxYAiQbSqUZNIL2gFohRfBzssoWrkIaWqVaxAsqbVeuOGV3zweIg9mGmoWYglceKUhfPGtws3AANue8NMh8N55g4UQ+4uXrRp1/pVOgcf4cKFzZjZ0c1JWKi0Hjk2eeeSHUPmUNASg4KHITv1hCg9jE+zbTMIduwARwt0AbMm58wy9dKUkzbddtxBoJk5AgRChKzp/VLM77N37FD/SauakG45dv48Hw9dCLjPHBvbFXnhL6Xtc2lh69YmRzcENvLbuHcvg8u6dj2EFBDsEKdYQeb4IsQ9bejXIANcWMOROTlFdNMF+sB34EC0XCEEhDacEOIJDHwiV0ZicHKSSbJcAAJ6GxIkRgZtgCjiCdG8oQWMA9mxYlp/gRAjRxlwAUY0JyR5gi8vlHAQCkuVlJVf+vA4pIxXfAKGkklGQsdTBM2SU1aAzcGOhlcaRIsKb7yhZBYviDMIUEtpZeccNxiY5kYpcKPNC1kEmkUzy4D5DzYk2MmVUpTs6dIgdAhyzKTH0EGGkwKJIckcbAF2xw5oOqrmMpFscMwAqGKATFgOfIMPW/43dIBPNaK6VMIyp6A6QBfIkDGJQLPMYQZhZtzRhpW1FiTGMo1s0MWzXZySikA5dNDWsN104ESyLqVCRiPQdgEDIv8c4dWwmu2Aj57capRCF3s8q0EXiOxzBDX4dHNbdO26tEk6ScyrwSaIlOPEDrd1I8QV/baUCiKIaCBxOLV8wYt4uCHQxg60NrwRJEnUIrHEBogCjxAZQ9dGIB5vNAkbS4yswRJW3JMDeQyS0rJGR4QzChNAM2FADNHhlx8DbYS68z+0MJEKG0APY4AFbdyjnw1II7uzGEyMwsowYBvQSdX5tXECPCdo3bIYw1Tw9QFS88MFAyGG+EYbhi7NYf7bnRzgNxb8+OKL2SHCk8WJehOkBhUVUHGA455YoY2bSr7xCcuJE2QJP40/fsA741yx5YhvNJk5QZBYQcPjVAzDhyeDvIFkki9ow8zpAykjAhZU9L7CO/IAQgeggXITyTK4C+QKDdJQ4QEVQ2BhwhFxBnpMFoLQkQLuFMSxiivPrxDGOGr8o4IgWUw6ABlkzHk6B1XQ4PwKVDwwjkCpRILqqbguo3S/YvDAA6qwAg+sYAUc8IdAJrEBMqCqCwNoxCYgkTlFTOAVBzxgGNphiUyR4RTh+tcm/pcsWkThAUUw4AGDMI7PlIBQ4YJBCL6gt1yM4xVRyKAciqAHgqBgE/4w6ILANrGHcOisZRRYQRCesYIylGEFqJCDhhzRCGQIUYiEgEEMSJgmWoyjCEFwYhmSIQd65MIgKECGyGS2BEKIQG1DEoMStpCAD4xxjOiQ4kEgEQNCjIwJbBiFMxzRLw58wB3tGGMyktEOTSgCRXBYAhs0ELQYeM0RcJyLGDgwBHf0YJGLrMMP8qE1UuxjFMOIGhOsUAELeIKLoFGCHNxBD1AmIwA/gAAFOJKIGBgAbMBkJRyssMsrUUAOOiDCL0AZgGTQYxfXcIkj4PBLv/mNHzTgnTxguRFa6GEbPSDCCG5Jzl9MIRvVEYUPDOC53lFBGeuQBj8Uwc2C0OIaH/7YwjncIYcANNOfmpCAJjLJod2JoHWte94ESkGDCbiiCcVsCQVyEYwtSECZ/rylP38h0HpyyBNVcMPznuc8D1ThFQ9QgCuGEA9FWEINqxEDLdRgCUXEow7b0IQ3iECPOmjUnwGQwBQ04VEZ+WMC63AFFejngaauoArvSAAsilCFOkQhGYoMhjra8YsfEKEFmggGUIEqg3PsIhsE1YgJXOGPMBhQhSrcxhbA6A4iJMCuRMgrLTWhjp+OVRO3gEA0h6SGMCgAFhMoYAYP+ESsbqMO6oisWMk5VqBCoAVTOEP50iQGC8aDA3I44CIVactF/rOylm2BBASb1p+gIBcfKGTCD0YQjBWMlpl+rawMALtaPRRVLN7cQjuC8ANNbGEbzBzrLWUAjSm0AKzQCMVvDyQGS2RjCxD4hTfO0QJv0OO7LpDAOW6BWWicIRsUmG6a1ACFXChBBxCIr3zPoAk9QGGzGwoIACH5BAkJAP8ALAAAAABAAEAAhwQRHRSK5ARQjYyOlExOVFSMoARvxAcvTFyu1CROZJTK9P///yQxOTyl5nSyxFSiyQw+YQQeNERwfjyKuARZn8zU2AR/4SSCxCxed2x5goyy1Nzr9wQwV3TA2nCgxFy76RQgJmyKnER/mSRzqTSU2wRAcARksKyvtymX6BxgkBRwuCQ/SRQvQDRfdAQWJBRRfFSh4v///wQiPBSAzP///1iv3czg8HTC5Dk/Sf///0yw6ZS/4fT1+BQ2TCeN2ECXygRyzFxufDR+p0SGrP///1SWsMTM1Eyl13Sz6SpIU////xxAVWy822yXuCR6tP///xQaHBRYiP///zRvjGSjuRU6VHvH2myizBQoMwRqvP///yqS34yitFRgbEl4hydjkHSCkOz0+f///xwwOzRqhAQoRjRYYwY2WyR0uCRZef///////2ywwAwmNKSyxHyQoFSEkP///wdqtLzFzhR6yDxoeFSq1hyGzKTC3EyZxGSYqMHb7wwpPwwSFQRWlAx2xKSkpESSuKzM6ERpdGyquFR6nIyarHS3x4y23LzW7JSnt9za3Hx+hP///3yozHTJ6QR41fr8/Eyp9Gy69CQ4Qf///+zu73ygvFSCqHyWrCqGyP///////4TK3FRqfFSSpBxGYWy20kSq7FyqzARep9Ta5HR+jFS27NTm9DRGVDyGrHy67Nzm9HSq2CRqnMTi/KjS9Hyu2FSW1P///yRJXUx4nDye7AQ5ZWy08P///1BWXBxqoBqG3Aw2TDSKyCxmhKy2xEyq4GzC5GSqxEyeyP///xyN5AxQiAxvukSk31ykwBQ/XAwgKwyA1yyBvpSuxHzA0XSMpAxipBxvrCw+Sf///wwYHxxQdFyg2ByByNTe6Xy+9Bw3RkyErFyYrHy06BxZgTxxhxwqL5SgqXyEjyxZcsTGyBx5vEyOrFx4kJS41gx6zBxmlKS6zGzC7Dx+nCx6tDxqlJSanAxenOTm5yxqj2Sx0SxPX9TW1HR2fOTu+mS87EyAlix0qQxCbP7+/gj+AP8JHEiwoMGB9E6Y6oIDC4gI1piNSdUFzAl6BzNq3Lgxkrl8DAAAsAaFGQiTzLBgEcdCHJ82K8AYicSxpk2BlgaE7GONpwuRP11EEBohgow2Y7iNYUHLkKWbUAtaYtSzTx+RQF1oHcq1qFEZMsqMWcGtypunUW1GkgfCKtasALYS9fq1jF27PVasWKaIZlqNi3AAuPo2q9y5X8GWCWuXw4FeSZa00Pb3IKDBhbFqFWoSyxhxY0IvPcDnLocyHFJzoHUPlJvKAyPlw5w5rjUQDpm1KQo2LB8W3PRy64Va9RkOEMyAiuY3LQ9dtDOjjDAmyLg5i8JEimRpkZFxQZb+HOCWhFYP1bfS3yqXoBAP5wSiv71tDYvMmxWiLauS5B6EM2eot145Xrx3UyTQEVYYCNaMMY6BUfGgCChLmEGLgCWUUE45tTTHUQbyidQHFBGQA2FlPGQCwT1m+JPhixhg0IRNlymIlTXWUFMBbAaVgsE1LYDyYobxgNMOR6X0lBkU1nRxIo8D8VAIKHVck6EAJQggATiUZRQJDjaKBIULGXgIJUGRNPHCIOBkmWUUXsRj5kDyhDmSNRmcyZGaErwgAJYC/FKPOgdZ0lZhzDDTxZx6oolJGhIc8+ef/LAThkFUFRYRNU82ahAP9ZARzqQCsPPOJVLhaBgWzOzoKZL+x3jxxZ9++MHPLpcONIA1WsXlQqLkvGrTJezwI0CtfuzyjgaxMWCNr0KBMEanwhoUSQoi1OMHBdwWAI9f5rjA62YnjVOtTep8UcC23ArRTyICZTDmZhFgwQK15xbEgzQFuMItKfxeIRAlIMhlEhj52uRBPwVQQMrDRcDzDz2/blWvDK4mvJEN0uhBigkfq+KEPicwY01X4nCjsU27fNIPyCagMYEgYDxUlAttYBHEyjXJMoQqJgSNTBGtdCFOV22wYC7PG6kDTxFBm5AFFbIkIc5cLLQxB9MbJTINFUFnkYUyvnzmlQxj8LEI1xptIAchcoidxQPOqJQYJWXgy3b+GG+rEHcWxFyQ9N1lMMr2P5GYQMg5WRhgQB7ZJN1bGSsUfvhBkWTBxjmOG0BMNiyMwZgMKxygN9dhIOMAHZ0/oEkVlDBG+QGsXG4QK8gcggwQvNvhwz1437VCD1vbTtAe2TjAO+/2kOBJ8HYtUYUixhO0gyb2GMA7JKHA8AY3BxTXAy3pVD8QNsQ8AAQkQKxjDxJucHPeAam1Zr5A2TS/PiQ+HAELPb1YQnE44J+M2c4GdIDGDNgHiWRsQR//uEcSjEMLWrzBfLj4gT0gwUHubUEgIVgBBDgAoOS84HQJiwQd7EGMDjZjH7gQiBGWsYTjpKccS3iG8RTgA2g0g4P+FmiALfYgkEgkwAwCAkULroHCakViBvY4AhAh0YEtNIcLtFiGekqwRFQd7hs+sAIvIGEBC2zhFDsgiCVAgcQS3KIEVDpGKdiGinV0IBlkJOM+tpCrgTThHi94Y4Yw0IJ6NFFPPOhfB/JoARR8YBUGCQOQhiSAQaQBE4bzVCRgsIVOGMMCZLSAMPh4kGfcIw1XKsELvBCFGWkMF3ewAh7LGEQdKMBL8ajDC9wkAHB4ARxNyCSPIoGEWNqBlhYIwCNskUltgGNUWfrTF/iRggIcEiqbNAY0ToFMCzwCBajgSDvAUQdASRMOruhHl/SECh+QwAo1ACUtP9AAWNikCb/+IAOpBFDNd8zDEdf0EhKaYQdZdlMHp4ihWjAxhSlMqlbzEEEB+rGLHQQ0NgrgxRY6cAMfdFMU+5CEMPXVjSmEY1vsokA/PlGAEaigFTa4iQ2QMAMf2KMTR+imBXTwAZE6pxv9EEFKHWYCIejBG0JQwTmwsYNEbIAH2+HBBvawgxpkYwa2YAI8eaHTD5zCp3+JhAfqUQBpOOxhHzPB0wihjEA4IxsqWF777kACYuDCCkwgRjN0aoxHiGISI+XIDlxRACE8LGomiJscnBEIKjjgEJA9BDQmGwo72OIOoQxlGUXxTXvqiRVCEII3RpBWqcnNAI37wwyywdofdpCRtESTwSN0YIsNvCoS6kBDIIqABtOe1nHLe+0UkYkCd3wABQoI7F/C4AjdUsEXKhBb54K7PyDSMgANeMRxt9HHfIUBEc7whTKU8QNNrCO4woVEM1AQDGEIowHIhSDXIrGHVmhCEz8YRSjsgQA7+NcOH2DCI/YRDBSQABevUG6+9CGIb8CABD7wwR0ivAUSwGAVsJAvlAICACH5BAkJAP8ALAAAAABAAEAAhwQRHRSK5ARQjYyOlExOVFSMoARvxAcvTFyu1CROZP///5TK9CQxOTyl5lSiyXSyxAw+YURwfszU2AQeNDyKuARZnwR/4W91fCleeSSCxNzr9wQwV3CgxIyy1HTA2myOrFy87ER/mRQgJiR0rjSU2wRAcBRvtwRksRxgkCQ/SDRfdKyvtySW6BQvQFSh4gQWJBRRfCRQbBSBzAQiPMzg8PT1+HTC5BRkolSavFiv3Uyw6TlBSRQ2TDyWzCeN2ARyzMHL1FxufDR+p4zC9ESGrP///////1SXsUyl1////3ez5xxAVRU5VGyq3G6Zumy82yR6tP///yRIWRQaHBRYiP///zRvjGSjuSZkkevz+X7I2v///4yitFRgbEx3nxQoM////wRqvDRqhBwwOyRZeQQoRiqS3jRYYwY2Wyx0qRR5xFSp1sHb7////2ywwEl4hwwmNKSyxHyNnFSEkP///zxoeByGzP///0ycyGSYqHyBhgwpPwwSFQRWlAx2xFyqzL/FyqSkpESSuKTC3KzM6ERpdGyquFR6nIyarHS3x4y23LzW7JSnt9za3P///3TI6AR41fr8/Eyp9Gy69CQ4Qf///+zu73ygvFSCqXypz3yWrCqGyP///1RqfFSSpBxGYGy20kSq7NTa5ARepyxmhAhqtNTm9Bxqp1S27DRGVDyGrJTG7Hy67Nzm9CRqnMTi/KjS9Hyu3CRCWFSW1Dye7AQ5Zf///2y29FBWXG96hBiG3Aw2TDSKyKy2xEyq4GzC5GSqxHyGkP///xyN5AxQiAxvukSk31ykwBQ/XAwgKwx/1yyBvpSuxHzA0XSKnBxwrAxipP///1yg2AwYHxxQdByAx9Te6Xy+9Bw4R0SXxJTB50yFrFyXqnSo0CxIUhxZgjxxh5SgqRwqLyxZchx5vEyOrFx4kJS41wx6zBxmlKS6zGzC7Dx+nCx6tDxqlJSanAxenOTm5yxqj2Sx0SxPX9TW1OTu+WS86kyAlgxCbCw+Sf///yxScBxknFyWxP7+/gj+AP8JHEiwoMGB8Vbo6eJtTItkcFowoddJTpx4BzNq3LgxEqBclF4kS/ZFnLgvLVqMoUQpBSUmPBI4AxKJo82bAi0NYDBtiohk05JNGDp0xowySA/wWOJNypIYXLLgnFrQkh4RL0RMe8F1wguiE4yWOZq0zIZdS85I+aRJKtWbkd6J4DNtGgAAXL+CNUoW6Ya/f2tt+DSOHzVmNd9qbLQDgN27eb1KLnq07wazG9BsqMX5kwoy7lopPhjoBR8+d/HinQwn4sqWKWIvYaIZDWfOJXKPU/FN3eiBkS4AQJ0ab1eSJ7/s2YO0zAGlTKTQoycFwu1auUvAKPTNSeK3NXD+DS8eOZk4iNiChANkrUakSJZaAWJk7hMEKWfOfMqeW0AJcGJsUwN4BIxXHF7J/NRCLhLgJMEH+XimAgzZ+SeAGO4QMSBOkYhHXGpf/TRGOBtSFQkz1FBTBwYWWmgFONt8x5FwH6aWYDLBlDhaDZfkU0chFAogpADgWMHBTYEYWJxW+jT4W0HWYEHGGygMKUAfIbgyCEeinHYgAMm80IWOTw5UAyZUUClkH31UUMApomUUyQ413jXFC7nIWCZBkTjxzRxVslnBDQWooudA73zp2DS57MmREyjMAU+bFVQghBDnHGTJFHXS1cWhjvKJiTwFwFNpBaNwM4JbBOmhqAj++pAZqkGRPBNCCKiOMsoIRMRS1TR1TjGBk7NuZA08BaSB6gkn4LAqQQN8WVcwxd6UiSvcjHKCtiMIoghwDNQpQguyVmtQDc8UIASzzB6zTGKAKOlYMuGYe9M5I3DDbhgU8MKGQBfUmNUX5dpbUCQmcNPOCWGEUcwxTQgUbnEjNWqwTd2oggfDDQOzyT/xOAZZVsNebBMNpRhSTMMGZLOJPSuoltcXY5h80zNXLBOGAQZkkA0sudwZmQhB2GyTNIIIwrMB6PyhRBd3GjdSvUZvdA4vx+xswA+guKCPCKu98EUygFS90SLkuMHzDz/MQ4I4WHX1xQSNmK2RBn4k4gf+2z8g4EOCAHj1gjgzFGx3Fnmr8QMkP6zhQ1BdTTDGBKDa/U8kBiRSDeOQrGHHBCJMNsYMldsdyQ+J2AHJ6o6LEDpRoxtuNuLNyLA4JDn4MMbcsJdRt+UGaYBOM2qsjrsZqfA+1Bh7lA18QWxU04zxkPxCQhfiEDVDSlQ/P9AQJDyxuAWQPOKCM1/MENYMB2BjjvcESYPEGqtboMw9rKygnFh7pLAE/AOpxhN6QD4LmEEHsIhHRMRShhQcgFjPo4EutKALSFjAAr5ggT3+IQsGjAUp2MCGMwCohB48wYIWtJ5AcoGN5pRhF95AhuwuFgllPMEYFrRAANZxC4EAISX+LpQCExjhvVX4QAvKuKAFGtCAVwgkErJIgXP+wgRvfGKG5oqELuaRAyVa4BEs+A4isLELsxxgA/RAhiaAxwozUFCJLADBEAhiCR6kYANn3AAEzgABUdjNFMrwgC+8uA4WsEogzpDFLgCDBulgAIuOqoEP1mADL8bxGgbJAlMyc5szUOMQpQtVJFxAAi0Mw4tgPORAGLGET2zGNvlQwSc+YLNbDEMLDfBiKHSwADnVgR7Wwc0n6gCDD4TyN5G4RTW0sAYvDuMRtChdI/JxhtvkhhpAwgQkpzJKH3igi6hkgQY4EgfCYCc734gAGeRBg1mZwgelbKYXQRAKWNzkAwn+GAd/SkCFCLiDGJnYppyUgI41aIEWXrQAKlAxCQ55YRz6LIGFiAGON2ABBecQKHC0UQ0zeOAJPsjhBUMBAkkckyA1eCgGJCpRIWEBH/hIxw040M6b0EAJyvDBPLSABJFeUAclPWlBaoAPDNRBSC0VEqnmkIYbPMMf51iEBtwTiRpogA3akEY11NADNzTDFzJAoRJBgAqTjqZPZIgADKwkqDRsIw+eIEI7nlEKjhmgGNXYBB7m0Yx5GEMZ9VNiAB7RgFsIVSPn+MYb5DEkNlEKGrw6whUMQVk3POABiQCGA3oQ1tsVcIlgtKejWpEGMeADBYLKla60dYJSPIMc5FChw97YxrnAWoAFNtABLUxRrEh0wBXswMcpKrXafTVMa1ujrfHIx4J7zIMFCzgsVSxxiVOwQ13QYC2zwnCCpSX3dqtTRgO+yYJrqNJcWehAGtJQjiOoAgrF4Bhy+YYOHxjjCfNogBmGcF4asqEb7RgBBXBwhWM4AA8IdsAa5gGKeSDBDGZQwiukay57EOIBs+DFMqrB4U1sggQuUAIsNringAAAIfkECQkA/wAsAAAAAEAAQACHBAgOBIfvjIqMBEZ8////TFBQR42xBGasBCZDhMvhRJ7UJCYo////Z7PQKkhSxMbMLJ7kBjZbZMf5JIjMZ6Gx0eb5JHe2BBYjRG16bIecBFib////JFZ0FCcxJGaMbL/eJI3cbHB0R5LBFBkaRKTkBDll////pMrszNrkdNL47vDxBHjXKXapNFZcfLLc////EJL0F0lq////ZKzEBBktVrPjJDhCRHyU////GzA4rKqsBClMNJ7kGzhIdMfh////3N3g////BxAUKFiAGYfZh9ToVKzc////4efpcXuETJi+VKfsF4LUVLv0BG/DbLXsZJioBl6j////rNb4ODg8FD5UVIeXPIas////Bn7bBCA3////BEyHVF5odKbUVG6EPJLcBj5r////////////JF6G4O/8FGagV6HBKVBj////FHbEHCkxKJLhFCAm+vz8N5ja////OJ/wGY7ofHt/VJjQFHK0lJyhDDBEbLrcpKq0eLPmFE93ZK7M////BDBWrLrMvMbMPGZxVH6M////NFp0OHicv8LEJDA4xNLgpMLclMjwpNL8OJDI////PHCIXGZsLH60lLXOKW+kfL/4vNDcbKLMvNrxXI60pKKkVFdZSazuHGGSFEJkfIqXj5OUPEJEjMr8VI/BHJPpXJSkBHLNNIrH1ODwdK7c1NLURHKC////T2Z8GW6s////F37KNDI0LI7M////DB4k8vb5OF5sdLrMTIKcfM7k////hNr0cYKR////////PKb0////OmJ4J0lkKZ/2////VLXv////FInn////////BGKuBIHmBEJ1XJ7UNGqEXIKk////SKbcDD5cb3Z8LH6sebrprLK8bLr5fIKMlKKspLLElM78DIbkDEZsDCg7bMfvLInEDBgfTGqEDFqcLFdsLI7YrMbc1NfcfNT0bKy+LDhITHuMPKHffMjcXKzQPE5cXKrkXLzuDHG8HD9SDCEvfKXHXGp0LGCBHGedHHi5HCAkHFB0bKrcDDJMzNHUrM7s/v7+CP4A/wkcSLCgwYFIqu364q5HD378qgQbos4TICQHM2rcuPHNgyQ25nVjk6OHjR7yUjpw1yINh072oCR6w7GmTYEqPtkA56ZDB5FatCBAsOOP0T8losUYBwwYB3v0aN2cWlDFtVk8Z2mhQSOoUKJFjUYoUSJMmGUx7KlqxomeCqo339xxI2QEjQt4uwYdSvRP0Qhjy55dNoBLGXWPOEmiCVcjEFBCwAnBm9fr175HyZZdRrgwFy4ebhkyhLGxwUxCUk+mzHUrgm54UDpYmaZlmhjcOA/wzEWDBkMGJikyPfCNNACqU7PWMq9kDjx4+OH5AzhaJ33jBAmyx+dz794aOP5hYmGJMVxampALUc/aDZsO3Xp80XPoFK03b1QAqYTNGQcuaqnCyWe++ZYMKVeIYN5NtBSAnHqrXUADG/PgsUs/N6EAxRBlqHIDJwVqEEUUBoggglRxpQfAiqrh5cY8PWCDIlW0SMKJB1YYIo6IyfRogBJ1LLhRCCuyiJyLNHjyFnH/vEGPODdgckYUPR5wgIn72JRJkSymRsMI6aTCZEGnTMICKRZQeYATB6BhygkcmbNeketdAM4FXcw45kC0iDIJFJEks6YTTvQxQWkHvUEFl+rdmYSQewr0hiWtUBDJmgeU8go6cEAq0CeMHimENJFuZAk+FOBDKKEKKLDIQf4qjMBoauBo4mmpktYRiSXxEFpKKQ1MoKdAdISaWjrD4loQLZEogYYTv5bSRjt7VBUqAHdiqOxGFcSDzjfQrrBCHiDoKYCxQlyzbU0ufNPAr+KSY8Srki4QKjgdJLvuspuCUYq4K/jQBmOHoHvHvjUt8o0tAGcBzToVCETkrG7oizBBb7zSABziZmHMB08IZC+XqSVxcU2oKNDAClm0LPA/SKCr7ckaVfAKO1msoIwym/Bghg7GskFzTUzkAUfLykBADCPHkSwEJENz9I4RRmSx8zbeUKIinUIcHLVGi8Dxgc47n7MELLNecMjXGl0yRwI775wCD/k4LQQQbGdkBv4TRWyzcwDeCDMni3cumfeyWSQwhzIBKAOPMCRje8GtbNOiTAKjMB7A405LZjHbbyhTROYBbC4MOA/SeoHhh/O5TRFzNG56PuvVfiferRNkxuvGlA64MOnU7iU4a+c+0CVtFCF7ACn4oonwdoLjtfH/hMIDLpoHoMsmdCSH1zz1UC/QO8TAozkMKVCih/cXuJGD+P/M4cMmvgvTBCMxKycEDW7QMHPrFTDG6HzXBF+Y4R+IkAxl5uGGXVCPGpvwge+YJwyBSEOBeWEDHih3MloYY36+Q581BPKAyeyPKz/BRu4WAQFdTHATm4hYkxAxi7zQoDk9+BzCaDEHbxBjgv66EIZ5PgGOu7SGDd3wxOEowQPY1U8CoSCIChjIla4gwAbdQAHbKrANXNDPdykQxgEJsov+6UULeMhBMHSIK1q0oQa4mKAwJOANg6jAJ60Rykm+wMFIvWEJTWzD8sKoL2zMogN7IYoD8JABmj0BBEXwRfY20QRtZOQN7ugAAhLJDwfwIwN9NM0b9jAHdhAjezDQhS8oB4Ru5IAvROmBA6qgDjbC5Y9EYIf5sncOMXJED64cSliq4I5ODMEcuKoAOeDAjhr8rXQS8AUjbLKLNO7gmkaJRkvC4JYx0YIaWTACO0gQt8YRowkjtMkbWOGQsBilBC2oRQw4IAlbGoQWi/6YADls4QMQxI1xm5DAJkLJJ3bKQyyaiYEggKEPPlhCizapACpeMYE+sAMaOftn+TZhT4LQIhzycMA7yWIWDmAAA/YADSYkUQkk3OcNtEDCJRbBDHysoRF9sIURXkG2v8GjCQMVZQaqkIZoaIYzyzDMIwahDkN4YEojqhI+IiGCGdiiDwrgqbiUoTP5+cIaBNVINogaDMF0xjMeeMQtMAEFUkCBAhRABzoooIRGrAFeW9XZOlLQhmlGCggMLUQMdOMZDYBHA+I4QytaYYcDxGNQeN1qG7yxCTmMsVRvwIY+7AEMPvDmsCKikqAG5Su8tiEPNWiDNsJKFRXQY7MYKGFDdwxboKhaiU2lZZnYUkuJjhLHtR4ogyHUMQlO7IhHo3VCMqA1AZVtDAy9BV0iLOEBDxjiFgbAhAG2W6I6tGMGllDAN8ixh0uwdl1IUAQ9RDENCywWH1M1RR32cILLMikgACH5BAkJAP8ALAAAAABAAEAAhwQFCQSE6YyKjAxCZESy+UxUXP///wxipAQkP////4TK3ESazARlsiRDUSQmJGTK/CSCyAQyVMjO1Gy00GSgtARMhiR4uAQSGtzt+TxkcQSO/Eym1GnC6URtdAhyxAU0XHyEjBQkKyxSZCSK1Cqm/BgYHKzG3ARWlDByoJSqvAsaI2xubBSC1BdSfESt7ESSvDSb3zROVMTi/BRDX3TW/BQ1SvT19XfA1AR0z1TA/CRKXITY9DyVzHyy3GSuzBwsMlSz6gQqRFyasBRutHTM7jSDtBSP8ER6lBc+XBRXhxQ6VAQMFCFllDRijDSU2FSr2QSK9YyapP///0SCpAlqsTR9sQQVJuv1+////2SWpDRTXP///6yqrAUcMP///+Dg4AR83XTS/BSW/ApKd////yQ3Pmy18WeswhRShBiK5Bw0QXzA/FSy/FmiwVh8nCyi9NTU2ARSlCR+wThshFSg4GTG+SSU5qza/AVYnGx8h1eTqjyi4BpJb5Te9P///1y67Ky6zJC83LzK3KSkqP///////zw+RBxyrMDAxDSo+ZTL9zQyNDxyiMTa7KTQ8P///5S21IzK/CRWhNTk9He57FSKr5SWnBxqoCxumJSmtPr8/FSEkuzu7SROaXx6fP///zSKxER+mLze/FBaZCyKzP///////8zm/P///////3S6zP///9/m7CQ+RmS66AReqJTC7HyWrP///yid9v///0ul62y+/BSE4AQsUP///1yu7ARtwQQ8bVySvP///0xyfBQeItza3ITS6GzS/Gx2fGyu5ESKtFSm7Dx2mAyF5ky28gwkNSyGxxwmMEyt5EyWvDyc2Mze7BxFWSxLVIza8Vy143zN6jyErByP7Ex5hwwOEAx+1Cw6TCyW5HR6fzym8MzW3CRtnyxabLzS5AyO7AxWhHzS8ByW8gxSiSx+vGzG7FRSVBRln0ycyCxETAwzUQxOhCx6tAwUGuTs9CyO2RyDzzxKXHzY+wx2zAwsQxxqrBw9UTyS3Fyu3JSdqP7+/gj+AP8JHEiwoMGBrADFctNEUosWSZI0iVcplqN5BzNq3LhRkwQQnZBIEyEug8kmc7Id2bRpCrJD8Yo10sSxpk2BV6LoqCFtmg5pAwbw4jWmaIWjeNihuJalUpV0ga7cnFqQE4gaalqpefeh64ehQ48ejXMCD55Xrw4Zo/ACQg+pVG1q8lfjh5p8uPLi6gp2DFGxceKYRfuKARVmZ9pBgEUzrsZgMUL8yIegct4IXvvyAiz4bGEGDHbtIjVhAT8Mjg8OWhYiBIIuli9/ndGp5Jxf2VQeQYHpAGHDu6js8tBuwghHqQdq8tYlhIousCsHwRVh3zQtIvjwKTqmArwkTDD+hdKj59oh0KJ34cAxQhUMM43j2hilAhh06JUR4GrlTpoSPm6kMA4rNmiiiQ2sjANJL5dcco0QQqQTmmjrbePKE7XEd5MNBcijghVW3FdZVjUgAUIjN0Vzhj7ptEEBBOqtBwYY/VST4VSadKjCBSDel08ZNeyTiQ2OaRKIBYi1sZ6MYABho4YbESOPPEuAGGIXk+UDApTy9YBPGxPUg8OMM1YDhBk2DbIElTz2KJk74CRn0CTMgHIDKWSCEQAH3yC3UTBULnFBlSGGsMwoRMppkCZ0QHADD2QGkIw1s0yikSaGLKFplTw6lweXigqkSTEj3ADDjAEEcI45e4D6jyX+AGw6qBX25REqRxPQc4MdeqaqjDKKHMRJCbFuaoUKKozi6q2i0gGNAslsk2oAO8wCF0HeAFCspsdykyizl45QjSvTBvBGHWtUpU2sxV6AbJzgcjRJMgrAUK451hIkgLbbXiAPCPHaRAkM1qQKBRQk1BGJcg5ou0Sx8vzwbcAa2ZDGDS4YTO0bjSHCb7sXWEKxTYrAYA4UARysDAEyCLTCx5pqE8LEI2ekSRrWfJMyFGLQYItADfOr6QXe1GyTGS6oszMU97zxzxcfbyuB0TWdgg015OycQyJXcBH1Eto0Q7VN2FiTyMFQJJLDHdnCvMQoY9dkBhB/HKxBz2us8/X+EiLHvVEk35iz9A7KLLI3In5vJIodO6DN9DfECq3pF4lrhIERfZBj9zAkrAszAJxUntEV5FBzjt0PkBB1rBcsm7gmAZiuuQZ1qL63637DbrrdtXsueeiiG2QDFLtDoUHqkUtOefAFXd6HEZu/YbjbiDNP0OJYo31PInq73bf1AkWyh+DGQ7EDG20LfQHc4AtkZg5o4+111Nr80L5AZbuANglrQy35oFMD3ynSYDqtJQI1QROaNooGPkp8o2Boo4HT/vEyycmMZq9LAxEydjBy+EwgHvtfyKynCDsIwwg7IwDLRJXATUUMg2O7GQeqYTCmzSI+IJDcmi4AMNFRohv+JzTYLBRGEE6s62FDew4cKncKejljWlAgQr4I4gmHuVAF3vKbDejxhBuUa4jpqooKtiUoZCkrhsdwgjV4Na0wTLEgsJIVj4ChAltRrRj1UAA0ehUAFwDLZpkqlhUu0JwufGpko6rHDZyBqgAYIQyteow8BjWrY/1ABR2AoZxsQAdFPiFSASBCN1CzETXNqk1YCsE04HUrOpFCFXQYE6r+0CebEMNfhLpPlmKhyQ31wAPPUAUPZDkjZ5zpJppYh7+sdB81lCEfQuplxY4khzOcAQJLmtEGnjQVDnmImdD5USvegYRYBLAm0SiGPg7RhjO8IEay3AAQbkSV+SDrPtH+QUAN3NEAJbwDQCkQBCuuYCAEKagX+jhAFfRAgWcMIT0yeoIzaiHNSxGjC/bBj3RwUQMdaEELnejEAPwCmCSEAxmVyIIeikCFCcWIBROAQQ9wx5HVGCqf+snLB2hTkg78wqfZYMkmQhEOdqAFPelxgnH8FCrILOMHy4gNdfgylJG2oBxJQEdgTvAb9DCgHj4wjaXANZe63CWnEcDMV4ZC0rGcoCyEaakF2tCGxdDUMVapQQ1aUYMI7GWtbBVLBQIzmFdQoQhteIEc3jI2TuikBg2YhjRmAJawbGawcXjFJRaqh3hA5Vpx8whIlDASLWRAHE1I7RyQEYqhvgQFZ5gWSfvmoZBsNOSqEGECCipxBhNgJFQBAQAh+QQJCQD/ACwAAAAAQABAAIcECA7///+MiowESH////9MUVEMaKxEi68EJ0OExtQrSFFEoNgkJigEaLgonvfExsxqssgHNltkyPz////R5vkkdrcLFxsEUI9JcXlviZwsV2xkoLIEdsxstewkjNkUJzMUGRpsc3c0Z4IEOGVQlLKkyuz///9MquwEedfM2uTu8PEEXKQ5WWEkl+kEHjQcUmz///900vhkwfP///8UlPQkN0QUNkksdan///////9Eeoysqqzc3eD///90wdlnq8L///8UP1n///+E0+p0x+cFftwcYJIHDxQUSGzh5ukLHypZs+AkUGlUrtw3k8xkmqwmWH1xe4Ss1vj///84ODz///8YcrR0tsx4s+ZcrOg5mdkUgdccaJ40hrxUboR8pcc4ktj///8GP2v///9UmsD////////g7/xXvfRUXmhUnsAEcMcsfrT///8cd7cUICZclaRMrun6/Pz///80XHQkQUz///8pYIJ8e3+Um59sutYqkt8cKTGkqrR8v/Y8dZGsusy8xsxUi6BUgoz///////84frS/wsQkMTf///+kwtz///////90rtxcZmyUtc680NwUVoxso8y82vE0isekoqT///9UWFgUQmR8ipePk5Q8QkRsqtzU4PDU0tREdoxkuuQXWoREmsz///8ajuKE2vcZTnFPZnxUpsz///80MjQIToQMLkRMptz////////y9vk8Xmz///88puz///////9xgpF0usz///8kbqQ8jrz///84YniUzvz///8nSWT///9EcIgZVYJMk88UieX///////////////8EQXX///8YesgMjvAsjsw2pvaEzuSssrxcntRcmqxQrvh8goyUoqykssQ8epxcgqRsuvxUmNAMRnQUZ6EMKDlMoMxsx+4sjtw8Z3esxtxUqO7U19wMXJz///981fNswugckuQsOEgcN0U0c5w8Tlz///9MfIx8wNQcMTocPlGM1OZ8yOQcSGQsUGBcq898stwkaJg8h7JcanQMcMAcICQsQkz+/v4I/gD/CRxIsKDBgUnEfRlmqEIuN24qsAGjDUuJMwczaty4UQ4kSfnysdMBpySJk2p+/IBwT1QzcFgmyeFIs6ZAFV9E3Ak26E8uLiuCrmhAlOiaZU5Q3QKlZU8CWDajFsRpilcwEaEuaL0gNGjRBmvCht3SBF6WPX5USLUp51E9DeGgDJi7lSvXoUOLiuXAAQUKLUSWtOg1c61GHrx+xXqRLNncAXW74tUblq9fFEW0PDsxC6Nhg9Yw2WMiZkTjx5AjcRkpCI60DRvUkNFFaVnly0VyLyHSQsrngXIyBGmHKYJpMadX3dHRkx0Xbgaic7PCRhcZCBC8bVnjN3fuFudm/mUrvBaWF3kKIhgfMQL5NhEYsBqB8whSEhVy5Kg4M+lRtGVsiHKLHnt0lxszRRCBxjhQSQXLKevUwYp66yVDR1yRSJKCTZ1wsgUlehCI2YFFoCHDCQ3WJAeE60yIgHoj1BMLE8I8otZacvSyhRP0NOFdEQjKIMM05HFkCys1IKBkBAiMYI8923xx429y4LMFKLyRyAwaaGRTUx/dJIkAKy9GYE89UJDzm0EUgKOZA1p+44xvG/HQTTwuKLmkAkF4keKaBMEyzh7PzKJlKQ54dpAc7fDRTZ5L+mNDBkUCSpAcWBDqzIHM0FCKM5UOlIcSH7gAKQJ1TGopRx0QCieC/sxMg0YvB6nwQamm5rmODV6EuuqlnD0DKzPMnOPAnwLZ8oYSuSLwwTq/IPurQbDssYQMQALpgAR+TPXGGy5YkGuLnkxLEwXEDLEprDEcS1AmSlggrqm3ZmJuTX5occ6wznALHCJvHOGCwC7Es46091JLChEnwFoEooU9IK/A4n6gRDUJ19SLA8ICGWsckwgUggVHHCGuBR+w4mvGBcFCDD3OwEpDDF7+gwjJAwv8hi0s19TBCeh4zEwMzvyThMkliwtuuT1zRAEppTiMhjNn7FDy1Ue8EU/TNZECs8cOoCFFFFiLq8Q+XNOUhazZzuzHJUcAcLW8eaTNUS+zECH0/jnTUBH31QMHYvdGk+wxBIlEByx30i7wMLhGZ6RLLJDfOFDy4khP+XjLRQxBSrYyOADA6HNbgPDmcnROA+iix006yac/nvoQe2SLhuWjL06y5psPBEvnnwfpQD+5x02y470XdMYWQxBzYOiq5I75IckXNAkpCWRbRLuXSH913dUPlC8RIxZxzjhRSL+4I+EPlAUqS3CKjh87qF8yH+0LtIUeWnjXwjRSSIL9Ssa06lFgCwlwXm6mkah/MGCAUWgfJ7SgBwN9QwsiUx8ALPCG2DVNDluAgCgMhI4OCOQQGiwZ+HqniC3AAzezmEXI/iEH4nnPAv3wYMZgsYUfeMMy/ijwwR7II4AUHoEavcMHG66AAst4oAndGogKUkiyDQ2uE/z4gShu0wEPIAsPVFSHDn8FiwD9YC+kaAIWDKIC1xWPZGlY2bTkoA02SMINYlkDBDzAO4FogopHCAHXJGGFDXRhDWBZgxMWoIiMyIEKA5RXCOS4JjlIwgAbOIBRjgIBMMiRHAM8QrzSMEYHDYMLT9CkXu6xBUVlpBLeu1q81GHFVXXiBvl4gj4m0wA1UKIENQlBLEsWLxdkopQagcUXIsEOOLDDK0TRhS44wZYCFE93SniDBdZRDWQG6hFGMMIgBMGNrjSgCyQQBSVbZk3MMe5bSmCFLQrIkRQ8QRjC/vjEIHJxF6/o4wDDWCe1rIk1ximBD3wolTy8UI1A8AAW+YEFDyBRDWwIYxtQCMY7RBCJrQjlAPoYhjcvJcyymSxcLmAFImoQDxuw4qURmNAIMGEKDfACA+G4wyogo5UVXKAcgriBJAS6kWqQDHAUy1U3kFSHOijgqfZggVSZ8IJtOIYuW7mBIHLRyFWRYxMFHViuTKWnCVGIPaZpzFUvsAoj6IAdhqCAueSQB21OzFTzQsCpXmRW9iDnqgOIxB+c8wiirkUFtghXMceq1yWVaT3taUxGRWAEKXFNBZqoAQfB9ai9vqhJxtkGE+ggghfcwbKDk8MDoiAPUn3AYOtYKIc8ZvuLXzyJCUiAwhM8YdiMJQEamThFO+QBUxvUgw5eyIQikrCqgAAAIfkECQkA/wAsAAAAAEAAQACHBBEdFIrkBFCOjI6UTE5UVIyhDG+6BC5MVK7cJE5cJG6clMr0JDE6XKC4bMTmPKbpBEBwzNTYBB4z////BH/hbK7EJF58BFmeR298JIXH3Ov3bHV8jLLUBDBXcKDENJXZbLPpFCAmbI6sBHDHFEBbJ1FvJHCsKD9IBGKsRH6crK+3OV5pBBYkGVqJVKHiFIHRBCI8////FC9AzODw9PX4FE54////OUFJGWSeXLrr////RJjINIColL/hFjlU////wcvU////bKrcOZzbFEdwFBocGzdIHHe4TK/ncrvRBjdb////BnbLJEdY////FIbc////J43ZWJi0ZLHRM1djNHGUDEZsTI2uXG586/P5////JFh4////SoSuNGqEjKK0JGaUBGW0VKjWJJLcHHCs////TKPUDCc0FCgxLHipPGh2BChGfLvuwdvvBDll////VGFsWLbspLLEfI2cDBgfHIHNHDE6FFaJfM3n////PHqYDBIV////////dMrsDB8rv8XKpKSkpMLcrMzoRGhwSneIbKm6VHqcZJ6sfLDgjJqsjLbcvNbslKe33NrcVISUfIGGVH+nRJ7U+vz8TKj0JDhBLGKU7O7vfKC8fKnPfJasZJikNFpsVKjwVGp8////bLbQ////FGq0lMbsfL7Q1NrkBF6mDGqsTIacHGKU1Ob0HFZ8NEZUTJ7MPIas3Ob0////////xOL8qNL0VIq0VLLn////JEJW////cLr0////VJbUOZ7n////////J3m3SXaZDDZMDIbkb3qENIrMrLbEZLbcTKrgfLbsfIaUDFCILFBcdMXkRKLcDD9nLIXElK7EPJbUdLLndIqcHEJZLHKsDGWnTH+PXKDY1N7pHFB2ZLvhTJq8PICgdKjQHEdnLEdPPG+JLFlxlKCpLGiMXKbkHCovDCk+XHiQlLjWTFZcpLrMHIbUPGqULJLclJqcdLbEHGqmDF6c5ObnXIycFHG5DDFKXLDaLHCY1NbUDIDedLDALF965O75DHLE/v7+CP4A/wkcSLCgwYH9BnnbRexZhjp1MkSJ5kLarH4HM2rcuHFSGw+/frlqYKhkAzENpoBKkmSKGXfupMmaxLGmTYFZOKRJUwDRlWdkwowQOqJoUSbshohhloSXu1FZbkotmCUTmG70XJExhQJFmK9Dwx5lwoQCUiR4EIxhE3WqzUnpwOh5VO2CXa5ev4I1OpYsBbO8mOVwt4CmW42vqni5Rs7uXVOQ8+oVO4Ks5b9/eeFBwkvD4YPrWhSqokyAAMd3rYm8QrKC6ykNJBF7YdkvZgR+xsz6PHCSJtGrTAu/IMAaTylXePw6YqB5vSPPoomBl8QMu7JlMY/B8yCXYbeTgP7xU1MagnABOFIU0BNPgZR0jDRkmTQpi4Y2PbBBlMTG2AfsmAnjQC2UfHcTDeLxY4V5wt2hB2M4eDDDTTNI88IHoGwTBWaY5ZBDgVKFJ84KEJhXogAWQJgODYfRMEodQ5Aihj4cUuAhiDaJsAwnJZ4ogBrh3IGJgS2y8cI9SWzIYQ5x5GKTHETw2CMEVmDAjyUT8laQKlFIwswHHAr4wG4bZUMCFc644UaJyhCyShcsamkQDS64w8wQYebmWUaTqLEMCUqoSeUK2ohApJy9SRPFnRyO4ccDhwrUCDVNdKCmmoSKgChHIESBh5IUCIMEEgscdAk14FhqqRvilHBIpP6b9ubCEMw8EaAf7rRF0BzU+NDBr258s4w4ccaq0SRR3LMNh7zk4CRBl/hwwq/A/lmKsTWpog8peGKWm67/KGKED2sc8Cs432iCrU1sfMAMs3Hg0dstJ5S7RgfBLPNNsetqRAM7STTDIR7uGAbEOUasUW4Ht/jQSL82jWInh5Q8IItAwxhxjsLlgkMCvxDz+UISYP4VgAMgCHRLJTBw7IMR04RskzSSGJMdBcy48888Z8gAQ8swnGBPBDJnqw8zL5hFARIfZCEHGmf83PIJJBRtUwaggFmWO0jMkgwaMEjwsz1GYGF1TeWIYUZ2wmzDBhzmiC22DDKMczZHo3xgc/52zLgATtwSzH0OIHdv1AY7pGDHhAMf2IFG4IHbcc4rhWukAROk6DOCWXG4E0IIErAQOQwgVz5QFv6QQhtZ90QhAeiiS2AHDLBWPskI8FxHVidj0PGH6KLbIUHthd+eRB2VMWFGFH+EwALwcZdu+j+ow/NC8p1EYc7vz0vwuCPTG6SBP/DUw0RlCERxAx3PP4/GH4SHTxAjdezD1xTRwMF++zCEYLf8A+nBMwzBF1C4YBhFAED7/vAHswFQINi4wg6IwgQxJEIFz1Og++zwQIEcoQHPCEsGuDGLeQBgDxpkAeiIBsAZnMIQ9SBKNJ6BEQbQAQAa9N0wHugNHjQALP5haAAxBLIBFOLweSFAg/TONglRSIEHevFHA4QgEECcEIcKDAEd/je9dJABEQYAiyue0QaBTIIBRtRgCOywRJnRgAxX6EJXvgLC7wwAi0ekAx2SMb1MuMcaXUHBEXbAAWjtIY1ZlEA+KjcDeawnMiiQAhnA9Q9I4BGHejxBG7FFA3zw5AKQMYUoruANg1yiCIgEgBbhQLxNTYIWYKAHDh5TgCNQUiDvuCQOnbfDoolAHo9ojGPSwIN08OkGqVQhC4bRys9MwgN3uEYVTGMXR7qidvlIZRZZAIdNtigSdyhEOIZzAVTEg3IbCYQus/iHEywyVqUgxyoK4YXzCEAPYP4QhE02sE4VMjAZ3uQIDTABAX5gYAsCYJAA8FEFDxyIANoEQPP+YIdxBHRO0NBGDdSghhqYyDTkCEcXmjkQGhCgnyxAQwj+IANIsLAmERABEaJEiC30yDQQIEcVIkFSgphUm89r3vtgULZxAOIVl6APDRwBhEag4xvBosIKluGMKZnHC/yA02dowM9kiu4PaDCHOdCwsZZx7AD28AGqqLCMJghqTWxSwyoMJadAALV9EgCrHexQib6e4ATgAEcTqEGNYFgqUHCFwCrUcId1xCofyExj+2IHORhEbWNnpdaq1OQMTmyhHeAzFg3ekUD2pbB7lTWrwg5wL80qwRnD0mgGNHo6lUtAQot6xCtlxaba1f5KCSRYRgLSdQmrXWIADGCB7/4QuuYGTmocS2sTBlsCTFw0VpMAxDAq4bskouG7aKCbHYxQCSMYIV/TAAJt+zUPOUACDrdgwDk2JgMjLMMTc1jHPDYVEAAh+QQJCQD/ACwAAAAAQABAAIcEER0MiuQEToSMjpRMTlRMiqQEb8NMsuwGL0wkTmQnb5yUyvQkMThMn8xstM4MP2EEHzMEXqQEf+HM1NgyjckkXoBsorREcYFsxuwkgMBvdXwsmuSMstRsjKTf7PgEMFdywNlUoeIUICb///////9EgpwEQHGsr7c0UF4kQEwEFyQsdagcYJA8YGpss+osjdlMkrxhtNgUL0Jcqs8bgMkEcszM4PAEIjyUv+H09fg5QUk0fqw8m9lcu+pamK4sV2psq7tsl7gXOlQUOEwcTm/EzNRcorwcQFdcbnwYKCsUGhwMZqQHN1n///////////////8qSFP///8ahtj///98yN50gY8aeLr///84dqQ0aoh8tec8pOckl+qMorQ0cZE8jLgkZozr9Pn///8EeNZcipx8usxEepR8jJgEWJw0WGQkerQZZ6AWgNcEKEYaWIAWKTf///8MEhVsutYkhsykssR8wdQESH+8xc4MGB86Z3lctegcMD2kwtx0qtjB2+9UYWykpKQijdqszOhEanRUepx8pshcpOBclsSMmqySuNm81uyUp7f////c2txUgpT///9cntRJeIp90Otcrtz6/PwkOEAsYpTs7u88nuxklqR8l698vvRUg6xUkqT///9Mptz///////9EkrwcRmAMdsRMptQMJjTU2uRMhpx0uvTU5vQ0RlRkprzc5vR8hpTE4vwkkuSo0vRUanz///8kUG8EYa3///8kdbf///9UltT///8nWHb///////9QqvQkSGAUecf///8MhuRseoQMNkystsR8uu98kqSkutD///88XnwMcbxUoMZ0tcQMHygsX3h0x+QsgsiUrsR0ipxMgZf///90s+lkrszU3uk8gKREnNJkvOb///9knrR8gIZEptyUoKc8coxEirD///8Md8xMd58MWpQ8WmQMJzwMSXfExsh8rdRceJBMVlxckqQ0mtw8apQserB0uswshsKUmpzk5ugMT4FUjqJUs+osT18UP2AMYKAMgdzU1tT+/v4I/gD/CRxIsKDBgR4GXQtB4YWgh4Ji8QihSpaYgxgzatRY6Y8fehS2zZgHYl6Mk3NAVKnC7QC8DapgVdpIs6ZAMYqk0TPioAGFDOVqCCVDlGg/QZn2VMHAZQOOizajEhTDDtcoCzBoGNhqQKjXokQlSCAjQVC+ST26cIIqlWYlRbjGedthwFZdrluH1gBLVqxfLlV6bFgws21GV9p2aNphy67ju3i97gU71q8ELpMOZPJg+GAfNvjGNR7duC4zaaOatXIwr7WDGKbgTQlrWcKwfNE2yOo8sJIFBfgUpEkTgTQzbUaAwABDj0awUsGC0aDAYwYIEJRm15bQZRIXVYXb/lbq9KXaueER0keI623clSuRcCzyIKZSJTEe/uA4NIUOODtzwNOXbRIEgEE+v4RnUw4liHPGcBCmwc8478iDCxA22GTDNW3Ak1IX2/XQQ4JRjfeFOGkIEOEK72ShwFOGVbLAFDxUQYllw0ggIok1BQGNOHcIIKSK1VTDDztsdVbJFm3EEI0gtfWQjws1JfPGBUOqeE4hCqyQDW8GrSLINlXAUxtTu2mUzZXqBCnklmHgkwOYB+XwywtV8IBjNF1whlEl8ehBxB2E3nGPOSwEoSCdBFXigiBlWtZFNJksOtA0vPByhwmFXlBBEIxu5AKeUPp1wC8LHCQGES08YAKn/nfocUknlobaaAg8gNCGZXwmKdAm+hDx6qu86BHGnLZqlMMLe1BCFlkbxMBJQZiog86wJqijhzqoJEvTKv2AIKBfGMSSpBfAkMIEEyYwwUstm3hb0zHwgNBPWBtQMq1AldSixrrtmtACEbXKW1AONDjAQ1hkRBNLYUXsA8wHTHxgwg9ETGNwTQu8AEJREoDDAywCdRBFMRSvi446yG6sbMICEjUFN1T+A0wUH+TMBCnAdOByTdeIBNYcsfxjzxAp5JyzPg9M8DNNNrRhRj+TmfKCGHUMMYQbSqNAxNM1BeMABZO9AI4saFiCANduCAFMO2DTdEgDDQzVTwxbzJKC/ht8fzDEEeHEvREOFGAj2Tyg6GMJ33wfMQQegmv0R8J51RADBUJYcgPjKRTjSOQZecCMM6XU0NUML8jAx+Z8790y6AWJYQAQV3RlQDN0pCPDDbzfsHfBoFcyu1ZdNUDDKXD07sbiwEdeiS20c2U8HHBAAAHvi78O+0BiLAEEM3c1kIEl1VsPweKfb1+QB94z41gz0kSRhPkQyHAK5OoTtAgu3jxmBAWASIIKzAeHUwQufwNRxBr65xgj6OIbp4CACgaIPCQgcCCI0IZoGsMMH/jhBHl4xgAHyIchXFAgbPCBPEZzBTAMwh4qEOEIC1iEC2aDH5pgw2i0sQbOWEIE/hOc4ClEYIULGmIFBSBOY3ywBoFoQIYThAD1tOc8fqQiC8QpjicsIJB15CEPQVSBCE6RiPwpggVliAB6VrCCP/CLAXkAgArkCAER8IGKYMsBC86AIggVYAXhGUAcwyiCZ6Bhe4ZgwSPOMxw2aIMDBMHEFwFAyQmKAAL+AB0qBCCJL6hIRWdgg6++EcdKxlAEKcCjwXJwCS1IYkhpOMcZDGEQTIiglJVUwjMA0bxQjecNkniDmzgpyoPUg5LItKQKiAG2IKjjAppy0yUUUIc/6UAOyKwkEInRSyUhQx2EgEahBHCPM2QBeI7IAzazmQcgAkKVvMlBIdTRAmi8qlDx/njDlzQSDgCsM5m3TEEmbYUKXrDqB8Mi1DJYkIyaaCCbAAUjGuBZkxxsggm1QIewsMWLT9mkEgT4JztvyYdwUBQjOWAEEYTwAxQ8gF3sMgERoGGObg4kBwSAKECBCAcr1LAmE6DGPoYADBSoa10wrQUvatoWnPpTpwBoZyFVwAckhAMPjsiBfXLgCDwwoh1HQMAQohCFzqUMqbWoRSFO+qeHQjWIIkhCEp7xDPpdj28ykIElUpCCISBAaTq7GCk6YNOMBEIOIs0mXOEgVz44lg+WYABkZbA2xgF2H/rYRzVD5Q9LIBaqcoziACFQ1+ttjnV941oxgHGEZezTVjmoccdnIRpG0Zqvd7xjnBuKcYQUHIERhZUKJjSghNlWMrRBtN5oTbs5BFjCEkNABibAholXMMCf65xjbZd7PThMFhiJmK7z1qGB634xhCGs4xCpR70bAIMYRQiuwexxAisAQgdwmGo6+BAFQFDjBPYIVUAAACH5BAkJAP8ALAAAAABAAEAAhwQJDgSH74yKjARIgESy/ExRVP///0yGnAknO4zK/ESf1StIUSQmLP///8TGzCSe9Gy0yf///2TH+SqJyiR3t9Hm+WagsgQWIwRYmmyKnP///wwzSjRXXwd2y2yz6SiN2DRqhBQZG////290d////0mSyaTK7AQ8aRQnL8za5O7w8f///xdJav///yhYfmTA7yl2qQQZLSSX6////yQ4QXTS/P///1KUsjSe5qyqrNzd4BSH32SqxEh6jITY8lCs3P///3TI6DqTyihplFSn7Dc3OQcPFARvxHPA1wRdpv///xwzPv///zRafOTo6HSz6XGCkf///6zW+DeZ2f///////1Sz6BR4wv///xaC1AQgN0dmd1SHl1mfvCRPZzSGvHylx1RuhP////r8/BRPeP///yRBTiRupFy65JScoxxmlwQuTHyy3Dh+tAweJBQgJ0SazBwuNASA5Hm/+RyQ5P///////xRwtODv/P///wwYHjJgcjyS3KSqtBwpMRRgnKy6zLzI0Dx2kFR+jHySrFReaL/CxITR5sTS4HSu3KTC3KTS/P///2qkzFSX1Cx+tDxwiJS1zjxCTCQyPLzS5HyHkXx6fFyWp7za8VyOtP///6SipP///zRxmJTO/I+TlFRYWFSOvFyrzgRyzXx/h9Tg8NTS1P///2S63P///2iatBh+zFimzBdurDx+nP///xw6SgQ2YAxenGa66ieS4v///296hAxCaPL2+P///zym6f///0yCnIre9////3m67P///3S6zP///0SKsChIZP///////////////1ye1ChikP///////zR6pFyCpP///wyO9CyOzDhigDym/KyyvJSirKSyxHyOnAxIeJTI8Eyi1Cyf9WzI8AxZlBQzRTxmdEyq9KzG3Aw8X9TX3GzC5////yw4SHzT8Tyg5ByI3GyrvDxOXHzJ5Fym5AxuuHy05Fyy3Bx3uAwhLixRYTyItFxqdBxRfCxCTCxxoRwgJEyWwAx+3BxhlMzR1KzO7P7+/gj+AP8JHEiwoMGBeKTMIbINx4OHD7ZNszJnEZ6DGDNq1DgG0y8ZOKxI8OHjUBBu3GqQ9FEDzbRtczCN2UizpkBcc6agI3doli4ZO+QEGEo0ALQHBCT0qkFgWwJcNqMWxPWOFjx2P2jJ2cpVaACvRYc+eNFLwoM5UKXWHJPtww8kcPZ1nbv1a9ii0Kb5kLDN00y1GfHwgQO3gxzDiOnOvTsUWgACPlxeBGzQxCoIrDpo3rx53xRt8IKwO0T63AsruugwNoqmxrZFlAeOSTQBwoQjo3J3GNVhHzok7ODpmkJnR5YAO+jg0DXr0DkrDxg/8EFg1l+1uByV4HHniPfco9L+iUKibUKWdtkw4cE1ZgwuPJiywaNDy4rJbXehtQSXNiouffp04d2AR1zBCgR8rJJIBTZV4AE06LDDzjZgDfWCBPxFNYYjw+iThDvuJOEdPRAIsYoi/fnnCR3oHIKGY0Wh8QI4123UCD03iCjiEe500cUd76QIGE47vMCOVnYF8AIas9SkCAyXfChlK13Q80gpsRlUAS26HIKDUF5xMw1sGlVwxiWyJKFmEndY8IgjQmZ5ExG0HIIOV1+dI8NkB43RBhdnrJlEK5dQ0EiNchI0xhO0sPMlVzIEgQ6iA0XSiSsYYLDmJWc0kuhGHtSpFVfggJPNQSqowUWmrB7QTCj+lH6qKBERysUVOzLECYYgQwyQ6QD48HJGnLJO9QE8qBy2FQ7w/FKQCvz04Kuv3XDxB4PFblTKPkhMoWwH5NCSIhggKDPAtLyCkW1NvwiBRGL1nSrQGMqEce653fTAD7HrTpUFBOgkhsQHfyHiwh73DgCJMpH0W1M2E7y7GTpTYCKQKnuQcS8ZPdjDr8ME4XKZEJtlgYoHArmwxS23DHCLCy6oAnJNTwghim6jeBDNP06wMM8AJ5wwwDdkpDAzTRVcEcxuvCkQDR6AeMFC0Cfc8o09R9e0ijq34TYBHP4QMs8tVLPggjNZ05QMgAN2IMo7PXAQy9wneOFFNWlvlM3+I6wQyIMjTawz99xSB5K3RpjEo86OSfAQDQsLrBGL5OucoMPhgbmjTojeOfIILMSsIfka84ijAuYYqZCEBXdwfkM8G5ghuuiCx4r6GKu3ouMwsMMy+xqR24457hbo/iE98XgDCwJraAH8Gh8fjksSlxifxDBnwEJD86LLjiXqBjkhC5prDkPBOks4r4UWsHhjOPgFUfJHJpqqmUkbW9CwvvPeeIM3/ASJxBAOwCoMHCAUlUABAva3hjiEAYAEyUQnBFHAA4CBGgiQx/4QMAlYQHAg0eoVBgaghmYowgnyQEEM9hcHBPTjgynoxiC64SsMDIEfF6GBClcYAxSgAAr+H1SFMiCRsB7gQyC2eIMWYsBELcTBG9Gb2RjIAAmE1RASMvuHA2LgBiYyEQXySAMAI2EPSGDjXi5QBiLmNYk3XOACXkTBEqLYL1zYYw9NaNm5FnadT+jhjYB8gxsqAT5C2GNlQBsAC/agroGowA16MMIFjBCDCyjRFJhLwS324IWgAU0aZDgdQaAgyVJewA0hMAcdP4ULF3iBA1Q7ATZcQIjw5SOSpoRkIYQnqzEMQhwcEEcsOWAPURYkDaU0gjIv8EdbZC0DG1gHCwYXCxZ4oWF9koQRAKDMZf5xBLyMzRgysIZ1mGFyg+NAE4Q3jkhys5tG+GMhVumfMGzgHrL+m10sFnCLy2lkE9wEgEC7GUlzvFBW4yCGN+7hu9+xABbWqMkIBErRgUbyAqSgJ0ZwcQ0ELIEG3mDe7GABiwzYZAwFqGhAlRlJP6RBowPBRTVo4NHtaWGBzVNeGMIZspSqdKDxNMIbbHFQmvQDCmD0xiRQsD6cakF5W4DpP3Dh059uE55+qMcnDKED9oxBBTowRBrqsYQYyMMPfmDq/vbXv6hSZgwT/SlQg7rMbgISjnp4AwreoEE4rnB/S0AAFHiqEWTKtaIEjaceFhuCELjhDX8EpBeZmMI19AGhRTgsRa8KT1NK8q6VXCEY1+HPXn5im5pVKWft+tk3PhYF1SBJrFpUYInUInal8HyjHh4LBWOCTAUCyIdtN5vbSE7iE77N2hgMUQ8GaPaq7zTCJEbgANk6TAc5GEEBihAC1ObVHKCAQh9Km6WAAAA7MGJrRXo5Y2lXNnVmU0JEVzFsWjk4YnBTejVFT1Mvbm5iRHBCSDJBNE5ONjFRSXU2QzdKNElOS3RHaFgwemlKbg=='

# Import packages
packagesloaded = False
importthread = threading.Thread(target = importPackages, daemon = True)
importthread.start()

# Waiting animation
time_between_frames = 100 # ms
timeout = 20_000 # ms
i = 0
while not packagesloaded and i < (timeout // time_between_frames):
    sg.PopupAnimated(ring_blue, background_color = 'white',
                     time_between_frames = time_between_frames, alpha_channel = 0.8)
    i += 1
    time.sleep(time_between_frames * 0.001)


# Scrape data, look for input worksheets, and show loading animation
datafound = False
conferencelistexists = False # default
mythread = threading.Thread(target = getDataAndInputFile, daemon = True)
mythread.start()

# Waiting animation
timeout = 30_000 # ms
i = 0
while not datafound and i < (timeout // time_between_frames):
    sg.PopupAnimated(ring_blue, background_color = 'white',
                     time_between_frames = time_between_frames, alpha_channel = 0.8)
    i += 1
    time.sleep(time_between_frames * 0.001)

stopanimation()
if not datafound:
    message = 'Timeout. Please try again. '
    print(message)
    sg.Popup(message)
    exit()

# Narrow down exact inputfile
fileknown = False
try:
    defaultvalue = os.path.join(mydir, potentiallist[0])
except IndexError:
    defaultvalue = ''
while not fileknown:
    inputfile = getfile(f'Do you want to add {len(outlist)} conferences to an .xlsx file?\n(Existing scraped worksheets will be overwritten.) ', defaultvalue = defaultvalue)
    if inputfile is not None:
        if os.path.isfile(inputfile):
            fileknown = True
        else:
            sg.Popup(f'{inputfile} not found.', title = 'File not found')
            defaultvalue = inputfile
            inputfile = inputfile.replace('/', '\\')
            if '\\' in inputfile:
                mydir = ('\\').join(inputfile.split('\\')[:-1])
    else:
        fileknown = True

if inputfile:
    if inputfile.endswith('.xlsm'):
        wb = openpyxl.load_workbook(filename = inputfile, data_only = False, keep_vba = True)
    else:
        wb = openpyxl.load_workbook(filename = inputfile, data_only = False)
    if 'Scraped from web' in wb.sheetnames:
        wb.remove(wb['Scraped from web'])
    ws = wb.create_sheet('Scraped from web')
    outputfile = inputfile
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Scraped from web'
    outputfile = 'Conference_List.xlsx'
ws.sheet_properties.tabColor = '00B050'

headers = ['Conference', 'Date', 'Location', 'Nominations', 'Confirmed Attendees (JH)', 'Confirmed Attendees (ST)', 'Comments', 'Registration fee', 'Weblink', 'Themes', 'Region']
for n, h in enumerate(headers):
    ws.cell(row = 1, column = n + 1).value = h
    ws.cell(row = 1, column = n + 1).font = Font(name = 'Arial', size = 10, bold = True)
    ws.cell(row = 1, column = n + 1).alignment = Alignment(wrapText = True)

for y, row in enumerate(outlist):
    for x, cell in enumerate(row):
        if x == 8:
            value = cell
            ws.cell(row = y + 2, column = x + 1).hyperlink = value
            ws.cell(row = y + 2, column = x + 1).font = Font(name = 'Arial', size = 10, color = '0000FF', underline = 'single')
        else:
            value = html.unescape(cell)
            ws.cell(row = y + 2, column = x + 1).font = Font(name = 'Arial', size = 10)
        ws.cell(row = y + 2, column = x + 1).value = value
        if x == 7:
            ws.cell(row = y + 2, column = x + 1).alignment = Alignment(wrapText = True) # Allow newlines in price

# Aesthetics
colwidths = [('A', 66), ('B', 14), ('C', 20), ('D', 17), ('E', 18), ('F', 18), ('G', 18), ('H', 22), ('I', 30), ('J', 57), ('K', 18)]
for col in colwidths:
    ws.column_dimensions[col[0]].width = col[1]
ws.freeze_panes = 'A2'

wb.active.views.sheetView[0].tabSelected = False
wb.active = ws
stopanimation() # Just to be sure

outputwritten = False
while not outputwritten:
    try:
        wb.save(outputfile)
    except PermissionError as error:
        errorthrown = permissionerrorpopup(error, outputfile)
        if errorthrown:
            message = 'Timeout. Please try again. '
            print(message)
            sg.Popup(message)
            exit()
    else:
        outputwritten = True
showreport()
